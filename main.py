import os
import csv
import io
import re
import json
import time
import httpx
import secrets
import threading
import imaplib
import smtplib
import uuid
import email as email_lib
from email.header import decode_header
from email.utils import parseaddr
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from fastapi import UploadFile, File, Form
from datetime import datetime, timedelta
from typing import Optional, List
from fastapi import FastAPI, Depends, Request, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import (
    create_engine, Column, Integer, String, Boolean, ForeignKey,
    DateTime, Text, text, func, and_, or_, case
)
from sqlalchemy.orm import declarative_base, sessionmaker, Session

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./crm.db")
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, connect_args=connect_args)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ─────────── MODELS ───────────

class PipelineDB(Base):
    __tablename__ = "pipelines"
    pipeline_id  = Column(Integer, primary_key=True, index=True)
    name         = Column(String, nullable=False)
    client_type  = Column(String, nullable=False, default="физ")
    description  = Column(String, nullable=True)
    stages       = Column(String, nullable=False)
    sort         = Column(Integer, default=0)

class ManagerDB(Base):
    __tablename__ = "managers"
    manager_id       = Column(Integer, primary_key=True, index=True)
    name             = Column(String, nullable=False)
    phone            = Column(String, nullable=True)
    email            = Column(String, nullable=True)
    role             = Column(String, default="manager")  # admin / manager / cleaner
    active           = Column(Boolean, default=True)
    last_seen        = Column(DateTime, nullable=True)   # для реального статуса онлайн/оффлайн
    telegram_chat_id = Column(String, nullable=True)     # привязанный telegram-чат клинера/менеджера
    login            = Column(String, unique=True, nullable=True)  # логин для входа
    password         = Column(String, nullable=True)               # пароль для входа (простое хранение)

class LeadDB(Base):
    __tablename__ = "leads"
    id             = Column(Integer, primary_key=True, index=True)
    name           = Column(String, nullable=False)
    phone          = Column(String, unique=True, index=True, nullable=False)
    status         = Column(String, default="Новый клиент")
    loss_reason_id = Column(Integer, ForeignKey("loss_reasons.reason_id"), nullable=True)
    telegram       = Column(String, nullable=True)
    viber          = Column(String, nullable=True)
    vk             = Column(String, nullable=True)
    whatsapp       = Column(String, nullable=True)
    email          = Column(String, nullable=True)
    pipeline_id    = Column(Integer, ForeignKey("pipelines.pipeline_id"), nullable=False, default=1)
    manager_id     = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)
    client_type    = Column(String, nullable=False, default="физ")
    source         = Column(String, nullable=True)
    company_name   = Column(String, nullable=True)  # для юр. лиц
    address        = Column(String, nullable=True)
    budget         = Column(String, nullable=True)
    responsible_id = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)  # amo-style "ответственный"
    comment        = Column(Text, nullable=True)       # комментарий к лиду
    taken_by       = Column(String, nullable=True)     # имя клинера, взявшего заявку
    taken_by_manager_id = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)  # реальный аккаунт клинера

class StageHistoryDB(Base):
    __tablename__ = "stage_history"
    id         = Column(Integer, primary_key=True, index=True)
    phone      = Column(String, nullable=False)
    old_status = Column(String, nullable=False)
    new_status = Column(String, nullable=False)
    changed_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    manager_id = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)

class LossReasonDB(Base):
    __tablename__ = "loss_reasons"
    reason_id = Column(Integer, primary_key=True, index=True)
    title     = Column(String, nullable=False)

class TaskDB(Base):
    __tablename__ = "tasks"
    task_id      = Column(Integer, primary_key=True, index=True)
    phone        = Column(String, ForeignKey("leads.phone"), nullable=False)
    description  = Column(String, nullable=False)
    deadline     = Column(Integer, nullable=False)
    is_completed = Column(Boolean, default=False)
    created_at   = Column(DateTime, default=datetime.utcnow)
    manager_id   = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)
    task_type    = Column(String, default="task")  # task / call / meeting / message

class NoteDB(Base):
    """Заметки/комментарии к лиду — критический слой, отсутствовавший в предыдущей версии."""
    __tablename__ = "notes"
    id         = Column(Integer, primary_key=True, index=True)
    phone      = Column(String, ForeignKey("leads.phone"), nullable=False, index=True)
    text       = Column(Text, nullable=False)
    note_type  = Column(String, default="note")  # note / call / whatsapp / email / system
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    manager_id = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)

class TagDB(Base):
    """Теги — гибкая разметка лидов."""
    __tablename__ = "tags"
    tag_id    = Column(Integer, primary_key=True, index=True)
    name      = Column(String, unique=True, nullable=False)
    color     = Column(String, default="#6D28D9")

class LeadTagDB(Base):
    """Связка M:M лидов и тегов."""
    __tablename__ = "lead_tags"
    lead_id = Column(Integer, ForeignKey("leads.id", ondelete="CASCADE"), primary_key=True)
    tag_id  = Column(Integer, ForeignKey("tags.tag_id", ondelete="CASCADE"), primary_key=True)

class CustomFieldDB(Base):
    """Кастомные поля — выбираются при создании лида, хранятся в JSON в лиде."""
    __tablename__ = "custom_fields"
    field_id      = Column(Integer, primary_key=True, index=True)
    pipeline_id   = Column(Integer, ForeignKey("pipelines.pipeline_id"), nullable=True)  # null = глобальное
    name          = Column(String, nullable=False)
    field_type    = Column(String, default="text")  # text / number / date / select / checkbox
    options       = Column(String, nullable=True)    # для select — "|"-разделённые значения
    required      = Column(Boolean, default=False)
    sort          = Column(Integer, default=0)

class LeadCustomValueDB(Base):
    __tablename__ = "lead_custom_values"
    id        = Column(Integer, primary_key=True, index=True)
    lead_id   = Column(Integer, ForeignKey("leads.id", ondelete="CASCADE"), nullable=False, index=True)
    field_id  = Column(Integer, ForeignKey("custom_fields.field_id", ondelete="CASCADE"), nullable=False)
    value     = Column(Text, nullable=True)

class ActivityDB(Base):
    """Полный лог событий по лиду, не только смена этапа."""
    __tablename__ = "activity"
    id         = Column(Integer, primary_key=True, index=True)
    phone      = Column(String, ForeignKey("leads.phone"), nullable=False, index=True)
    event_type = Column(String, nullable=False)  # stage_change / note_added / lead_created / task_added / tag_added / tag_removed / assigned / bulk_update
    payload    = Column(Text, nullable=True)     # JSON-детали события
    manager_id = Column(Integer, ForeignKey("managers.manager_id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)

class SettingDB(Base):
    __tablename__ = "settings"
    key   = Column(String, primary_key=True)
    value = Column(Text, nullable=True)

class TelegramNotifyDB(Base):
    """Отправленные клинерам telegram-уведомления по лиду — чтобы можно было
    отредактировать все копии сообщения, когда заявку кто-то забрал."""
    __tablename__ = "telegram_notify"
    id         = Column(Integer, primary_key=True, index=True)
    lead_id    = Column(Integer, ForeignKey("leads.id", ondelete="CASCADE"), nullable=False, index=True)
    chat_id    = Column(String, nullable=False)
    message_id = Column(Integer, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)

class EmailAccountDB(Base):
    """Корпоративный почтовый ящик, подключённый по IMAP — входящие письма
    автоматически превращаются в лиды (карточки клиентов)."""
    __tablename__ = "email_accounts"
    id            = Column(Integer, primary_key=True, index=True)
    name          = Column(String, nullable=False)          # имя ящика -> используется как имя лида
    email         = Column(String, nullable=False)           # используется как контакт лида
    password      = Column(String, nullable=True)            # пароль/пароль приложения от почты
    imap_host     = Column(String, nullable=False, default="imap.hoster.by")
    imap_port     = Column(Integer, nullable=False, default=993)
    imap_ssl      = Column(Boolean, default=True)
    smtp_host     = Column(String, nullable=True, default="smtp.hoster.by")
    smtp_port     = Column(Integer, nullable=True, default=465)
    smtp_ssl      = Column(Boolean, default=True)
    pipeline_id   = Column(Integer, ForeignKey("pipelines.pipeline_id"), nullable=False, default=2)
    client_type   = Column(String, default="юр")
    active        = Column(Boolean, default=False)
    last_uid      = Column(Integer, nullable=True, default=None)  # None = ещё ни разу не проверяли (baseline)
    last_checked  = Column(DateTime, nullable=True)
    last_error    = Column(Text, nullable=True)
    leads_created = Column(Integer, default=0)
    created_at    = Column(DateTime, default=datetime.utcnow)

class MailQueueDB(Base):
    """Очередь внутренней email-рассылки. Письма ставятся в очередь и
    отправляются фоновым воркером через SMTP выбранного корпоративного ящика."""
    __tablename__ = "mail_queue"
    id              = Column(Integer, primary_key=True, index=True)
    campaign_id     = Column(String, index=True, nullable=False)   # группировка писем одной рассылки
    campaign_name   = Column(String, nullable=True)
    email_account_id = Column(Integer, ForeignKey("email_accounts.id"), nullable=False)
    lead_id         = Column(Integer, ForeignKey("leads.id"), nullable=True)
    to_email        = Column(String, nullable=False)
    to_name         = Column(String, nullable=True)
    subject         = Column(String, nullable=False)
    body            = Column(Text, nullable=False)
    status          = Column(String, default="pending")   # pending / sent / failed
    attempts        = Column(Integer, default=0)
    error           = Column(Text, nullable=True)
    created_at      = Column(DateTime, default=datetime.utcnow)
    sent_at         = Column(DateTime, nullable=True)

# ─────────── DB INIT ───────────

Base.metadata.create_all(bind=engine)

def run_migrations():
    with engine.begin() as conn:
        try:
            existing = {row[1] for row in conn.execute(text("PRAGMA table_info(leads)"))}
        except Exception:
            existing = set()

        # Drop legacy columns from previous Nexus schema
        for col in ("budget_old",):
            if col in existing:
                try:
                    conn.execute(text(f"ALTER TABLE leads DROP COLUMN {col}"))
                except Exception:
                    pass

        additions = {
            "whatsapp":    "VARCHAR",
            "company_name":"VARCHAR",
            "address":     "VARCHAR",
            "budget":      "VARCHAR",
            "responsible_id": "INTEGER",
            "comment":     "TEXT",
            "taken_by":    "VARCHAR",
            "taken_by_manager_id": "INTEGER",
        }
        for col, decl in additions.items():
            if col not in existing:
                try:
                    conn.execute(text(f"ALTER TABLE leads ADD COLUMN {col} {decl}"))
                    print(f"[migrate] added leads.{col}")
                except Exception as e:
                    print(f"[migrate] add leads.{col}: {e}")

        # tg/vb/vk/pipeline_id/manager_id/client_type/source/loss_reason_id — keep legacy migrations safe
        for col in ("telegram", "viber", "vk"):
            if col not in existing:
                try:
                    conn.execute(text(f"ALTER TABLE leads ADD COLUMN {col} VARCHAR"))
                except Exception:
                    pass
        if "pipeline_id" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN pipeline_id INTEGER DEFAULT 1"))
            except Exception:
                pass
        if "manager_id" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN manager_id INTEGER"))
            except Exception:
                pass
        if "client_type" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN client_type VARCHAR DEFAULT 'физ'"))
            except Exception:
                pass
        if "source" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN source VARCHAR"))
            except Exception:
                pass
        if "loss_reason_id" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN loss_reason_id INTEGER"))
            except Exception:
                pass
        if "email" not in existing:
            try:
                conn.execute(text("ALTER TABLE leads ADD COLUMN email VARCHAR"))
            except Exception:
                pass

        # tasks extended columns
        try:
            task_cols = {row[1] for row in conn.execute(text("PRAGMA table_info(tasks)"))}
        except Exception:
            task_cols = set()
        if "created_at" not in task_cols:
            try:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN created_at DATETIME"))
            except Exception:
                pass
        if "manager_id" not in task_cols:
            try:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN manager_id INTEGER"))
            except Exception:
                pass
        if "task_type" not in task_cols:
            try:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN task_type VARCHAR DEFAULT 'task'"))
            except Exception:
                pass

        # Managers — role & active
        try:
            mgr_cols = {row[1] for row in conn.execute(text("PRAGMA table_info(managers)"))}
        except Exception:
            mgr_cols = set()
        if "role" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN role VARCHAR DEFAULT 'manager'"))
            except Exception:
                pass
        if "active" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN active BOOLEAN DEFAULT 1"))
            except Exception:
                pass
        if "last_seen" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN last_seen DATETIME"))
            except Exception:
                pass
        if "telegram_chat_id" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN telegram_chat_id VARCHAR"))
            except Exception:
                pass
        if "login" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN login VARCHAR"))
            except Exception:
                pass
        if "password" not in mgr_cols:
            try:
                conn.execute(text("ALTER TABLE managers ADD COLUMN password VARCHAR"))
            except Exception:
                pass

        # Добавляем этап "Новая заявка" в воронку юр. лиц, чтобы заявки
        # из почты сразу были видны на канбане (а не только в списке лидов)
        try:
            row = conn.execute(text("SELECT stages FROM pipelines WHERE pipeline_id=2")).fetchone()
            if row and row[0] and "Новая заявка" not in row[0].split("|"):
                conn.execute(text("UPDATE pipelines SET stages=:s WHERE pipeline_id=2"),
                             {"s": "Новая заявка|" + row[0]})
                print("[migrate] добавлен этап 'Новая заявка' в воронку 'Юридические лица'")
        except Exception as e:
            print(f"[migrate] pipeline stage update: {e}")

run_migrations()

# ─────────── СИД ───────────

DEFAULT_PIPELINES = [
    dict(pipeline_id=1, name="Физические лица — уборка квартир/домов", client_type="физ",
         description="Разовая и регулярная уборка жилья для частных клиентов",
         stages="Новый лид|В обработке|Связались|Выезд на объект|Взято в работу|Согласование даты|Работа выполняется|Завершено|Отменено", sort=1),
    dict(pipeline_id=2, name="Юридические лица — клининг офисов", client_type="юр",
         description="B2B продажи клининга офисов и коммерческих помещений",
         stages="Первый контакт|Квалификация заявки|Выезд на замер объекта|Коммерческое предложение отправлено|"
                "Переговоры / согласование условий|Договор подписан|Первая уборка выполнена|Регулярный контракт|Отказ", sort=2),
    dict(pipeline_id=3, name="Генеральная уборка", client_type="физ",
         description="Комплексная генеральная уборка квартир, домов, коттеджей",
         stages="Заявка|Оценка объема работ|Расчет стоимости|Согласование даты|Уборка выполнена|Оплата получена|Отказ", sort=3),
    dict(pipeline_id=4, name="Уборка после ремонта", client_type="физ",
         description="Клининг после строительных и ремонтных работ",
         stages="Заявка|Осмотр объекта|Расчет стоимости|Согласование даты|Работы выполнены|Оплата получена|Отказ", sort=4),
    dict(pipeline_id=5, name="Химчистка мебели и ковров", client_type="физ",
         description="Химчистка мягкой мебели, ковров, матрасов на выезде",
         stages="Заявка|Уточнение изделий и загрязнений|Расчет стоимости|Выезд мастера|Химчистка выполнена|Оплата получена|Отказ", sort=5),
    dict(pipeline_id=6, name="Мойка окон и фасадов", client_type="юр",
         description="Мойка окон, витражей и фасадов зданий, в т.ч. с альпинистами",
         stages="Заявка|Замер объекта|Коммерческое предложение|Согласование условий|Работы выполнены|Оплата получена|Отказ", sort=6),
    dict(pipeline_id=7, name="Абонементное обслуживание (регулярный клининг)", client_type="юр",
         description="Продажа регулярного обслуживания по подписке/абонементу",
         stages="Заявка|Презентация условий абонемента|Пробная уборка|Согласование графика|Договор подписан|"
                "Активный абонемент|Отказ", sort=7),
    dict(pipeline_id=8, name="Клининг после потопа/пожара (аварийные работы)", client_type="физ",
         description="Срочный клининг после ЧП: затопление, пожар, форс-мажор",
         stages="Экстренная заявка|Выезд специалиста|Оценка ущерба и объема работ|Расчет стоимости|"
                "Работы выполнены|Оплата получена|Отказ", sort=8),
]

DEFAULT_TAGS = [
    ("VIP", "#FBBF24"), ("Повторный", "#10B981"), ("Корпоративный", "#3B82F6"),
    ("Срочно", "#EF4444"), ("День рождения", "#F472B6"), ("Из рекламы", "#A78BFA"),
]

DEFAULT_REASONS = [
    (1, "Высокая цена"), (2, "Не подошел график"), (3, "Выбрал конкурента"),
    (4, "Не оставил контакт"), (5, "Не актуально"), (6, "Другое"),
]

def seed_data():
    with SessionLocal() as db:
        if db.query(PipelineDB).count() == 0:
            for p in DEFAULT_PIPELINES:
                db.add(PipelineDB(**p))
            db.commit()
            print("[seed] воронки продаж (8 шт.)")

        manager = db.query(ManagerDB).filter(ManagerDB.manager_id == 1).first()
        if not manager:
            db.add(ManagerDB(manager_id=1, name="Администратор", phone=None, email=None, role="admin", active=True))
            db.commit()
            print("[seed] создан администратор")

        if not db.query(ManagerDB).filter(ManagerDB.role == "manager").first():
            next_id = (db.query(func.max(ManagerDB.manager_id)).scalar() or 0) + 1
            db.add(ManagerDB(manager_id=next_id, name="Менеджер 1", phone=None, email=None, role="manager", active=True,
                              login="Manager1", password="ManagerCRM100"))
            db.commit()
            print("[seed] создан менеджер по умолчанию")

        if not db.query(ManagerDB).filter(ManagerDB.role == "cleaner").first():
            next_id = (db.query(func.max(ManagerDB.manager_id)).scalar() or 0) + 1
            db.add(ManagerDB(manager_id=next_id, name="Клинер 1", phone=None, email=None, role="cleaner", active=True,
                              login="Cleaner1", password="CleanerCRM100"))
            db.commit()
            print("[seed] создан клинер по умолчанию")

        # Простая авторизация: гарантируем, что у существующих аккаунтов менеджера/клинера
        # (созданных до появления логина/пароля) есть логин и пароль по умолчанию
        if not db.query(ManagerDB).filter(ManagerDB.login == "Manager1").first():
            m = db.query(ManagerDB).filter(ManagerDB.role == "manager", ManagerDB.login.is_(None)).first()
            if m:
                m.login = "Manager1"; m.password = "ManagerCRM100"
                db.commit()
                print("[seed] логин/пароль назначены существующему менеджеру")

        if not db.query(ManagerDB).filter(ManagerDB.login == "Cleaner1").first():
            c = db.query(ManagerDB).filter(ManagerDB.role == "cleaner", ManagerDB.login.is_(None)).first()
            if c:
                c.login = "Cleaner1"; c.password = "CleanerCRM100"
                db.commit()
                print("[seed] логин/пароль назначены существующему клинеру")

        if db.query(LossReasonDB).count() == 0:
            for r in DEFAULT_REASONS:
                db.add(LossReasonDB(reason_id=r[0], title=r[1]))
            db.commit()
            print("[seed] причины отказа")

        if db.query(TagDB).count() == 0:
            for name, color in DEFAULT_TAGS:
                db.add(TagDB(name=name, color=color))
            db.commit()
            print("[seed] теги по умолчанию")

        demo = db.query(LeadDB).filter(LeadDB.phone == "+70000000001").first()
        if not demo:
            demo = LeadDB(name="Демо клиент", phone="+70000000001", status="Новая заявка",
                          pipeline_id=1, manager_id=1, client_type="физ", source="Демо")
            db.add(demo)
            db.commit()
            db.refresh(demo)
            # Демо заметка и активность
            db.add(NoteDB(phone=demo.phone, text="Добро пожаловать в cleanfloorCRM! Это пример заметки — комментариев под лидом.", note_type="note", manager_id=1))
            db.add(ActivityDB(phone=demo.phone, event_type="lead_created", payload="Демо-лид для предпросмотра", manager_id=1))
            db.commit()
            print("[seed] демо-лид с заметкой")

        # Settings default
        if not db.get(SettingDB, "widget_token"):
            db.add(SettingDB(key="widget_token", value="cfcrm_" + secrets.token_urlsafe(24)))
            db.add(SettingDB(key="company_name", value="cleanfloorCRM"))
            db.commit()
            print("[seed] настройки по умолчанию")

        # Почтовый ящик по умолчанию — данные хостера уже проставлены,
        # пароль нужно ввести в интерфейсе (раздел «Настройки → Почта»)
        if db.query(EmailAccountDB).count() == 0:
            db.add(EmailAccountDB(
                name="zakaz@cleanfloor.by", email="zakaz@cleanfloor.by", password=None,
                imap_host="imap.hoster.by", imap_port=993, imap_ssl=True,
                smtp_host="smtp.hoster.by", smtp_port=465, smtp_ssl=True,
                pipeline_id=2, client_type="юр", active=False, last_uid=None,
            ))
            db.commit()
            print("[seed] почтовый ящик zakaz@cleanfloor.by добавлен (нужно указать пароль и включить)")

seed_data()

# ─────────── APP ───────────

app = FastAPI(title="cleanfloorCRM")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
def _auto_setup_telegram_webhook():
    """Если заданы TELEGRAM_BOT_TOKEN и PUBLIC_URL — автоматически подписываем бота на вебхук при старте."""
    tg_token = os.getenv("TELEGRAM_BOT_TOKEN", "")
    public_url = os.getenv("PUBLIC_URL", "")
    if not tg_token or not public_url:
        return
    try:
        httpx.post(
            f"https://api.telegram.org/bot{tg_token}/setWebhook",
            json={"url": public_url.rstrip("/") + "/telegram/webhook"}, timeout=8,
        )
        print("[TG] webhook настроен автоматически")
    except Exception as e:
        print(f"[TG] webhook auto-setup: {e}")

@app.on_event("startup")
def _start_email_poll_thread():
    """Запускаем IMAP IDLE-потоки для каждого активного ящика + watchdog."""
    # watchdog: раз в минуту проверяет, живы ли потоки, и перезапускает упавшие
    t_watchdog = threading.Thread(target=_email_poll_loop, daemon=True, name="imap-watchdog")
    t_watchdog.start()
    # немедленно поднимаем IDLE-потоки, не ждя первый тик watchdog
    try:
        _ensure_idle_threads()
    except Exception as e:
        print(f"[email-idle] ошибка при старте: {e}")
    print("[email-idle] IMAP IDLE запущен (push-режим)")

# ─────────── SCHEMAS ───────────

class Lead(BaseModel):
    name:           str
    phone:          str
    status:         Optional[str] = None
    loss_reason_id: Optional[int] = None
    telegram:       Optional[str] = None
    viber:          Optional[str] = None
    vk:             Optional[str] = None
    whatsapp:       Optional[str] = None
    email:          Optional[str] = None
    pipeline_id:    int = 1
    manager_id:     Optional[int] = None
    client_type:    str = "физ"
    source:         Optional[str] = None
    company_name:   Optional[str] = None
    address:        Optional[str] = None
    budget:         Optional[str] = None
    comment:        Optional[str] = None
    tag_ids:        Optional[List[int]] = []

class LeadUpdate(BaseModel):
    name:           Optional[str] = None
    phone:          Optional[str] = None
    telegram:       Optional[str] = None
    viber:          Optional[str] = None
    vk:             Optional[str] = None
    whatsapp:       Optional[str] = None
    email:          Optional[str] = None
    source:         Optional[str] = None
    client_type:    Optional[str] = None
    pipeline_id:    Optional[int] = None
    manager_id:     Optional[int] = None
    company_name:   Optional[str] = None
    address:        Optional[str] = None
    budget:         Optional[str] = None
    comment:        Optional[str] = None
    tag_ids:        Optional[List[int]] = None
    custom:         Optional[dict] = None

class Pipeline(BaseModel):
    pipeline_id: int
    name:        str
    client_type: str = "физ"
    description: Optional[str] = None
    stages:      str
    sort:        int = 0

class Manager(BaseModel):
    manager_id: int
    name:       str
    phone:      Optional[str] = None
    email:      Optional[str] = None
    role:       str = "manager"
    login:      Optional[str] = None
    password:   Optional[str] = None

class ManagerUpdate(BaseModel):
    name:    Optional[str] = None
    phone:   Optional[str] = None
    email:   Optional[str] = None
    role:    Optional[str] = None
    active:  Optional[bool] = None
    login:   Optional[str] = None
    password: Optional[str] = None

class LoginRequest(BaseModel):
    login:    str
    password: str

class StageHistory(BaseModel):
    phone:      str
    old_status: str
    new_status: str

class LossReason(BaseModel):
    reason_id: int
    title:     str

class Task(BaseModel):
    task_id:      int
    phone:        str
    description:  str
    deadline:     int
    is_completed: bool = False
    task_type:    str = "task"

class TaskUpdate(BaseModel):
    description:  Optional[str] = None
    deadline:     Optional[int] = None
    is_completed: Optional[bool] = None

class Note(BaseModel):
    phone:      str
    text:       str
    note_type:  str = "note"

class Tag(BaseModel):
    name:  str
    color: str = "#6D28D9"

class CustomField(BaseModel):
    pipeline_id: Optional[int] = None
    name:        str
    field_type:  str = "text"
    options:     Optional[str] = None
    required:    bool = False
    sort:        int = 0

class BulkUpdate(BaseModel):
    phones:        List[str]
    set_status:    Optional[str] = None
    set_manager:   Optional[int] = None
    set_pipeline:  Optional[int] = None
    add_tags:      Optional[List[int]] = None
    remove_tags:   Optional[List[int]] = None
    delete:        Optional[bool] = False

class EmailAccountIn(BaseModel):
    name:          str
    email:         str
    password:      Optional[str] = None
    imap_host:     str = "imap.hoster.by"
    imap_port:     int = 993
    imap_ssl:      bool = True
    smtp_host:     Optional[str] = "smtp.hoster.by"
    smtp_port:     Optional[int] = 465
    smtp_ssl:      Optional[bool] = True
    pipeline_id:   int = 2
    client_type:   str = "юр"
    active:        bool = True

class EmailAccountUpdate(BaseModel):
    name:          Optional[str] = None
    email:         Optional[str] = None
    password:      Optional[str] = None
    imap_host:     Optional[str] = None
    imap_port:     Optional[int] = None
    imap_ssl:      Optional[bool] = None
    smtp_host:     Optional[str] = None
    smtp_port:     Optional[int] = None
    smtp_ssl:      Optional[bool] = None
    pipeline_id:   Optional[int] = None
    client_type:   Optional[str] = None
    active:        Optional[bool] = None

class MailingCreate(BaseModel):
    email_account_id: int
    subject: str
    body: str                       # поддерживает {{name}} — подставится имя клиента
    lead_ids: Optional[List[int]] = None   # явный список лидов
    pipeline_id: Optional[int] = None      # либо все лиды из воронки с заполненным email
    tag_id: Optional[int] = None           # либо все лиды с тегом и заполненным email
    campaign_name: Optional[str] = None

def _clean(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

# ─────────── FCM ───────────

def send_fcm_push(name: str, phone: str):
    fcm_key = os.getenv("FCM_SERVER_KEY", "")
    if not fcm_key:
        return
    try:
        httpx.post(
            "https://fcm.googleapis.com/fcm/send",
            headers={"Authorization": f"key={fcm_key}", "Content-Type": "application/json"},
            json={"to": "/topics/managers", "data": {"title": "Пришла новая заявка!", "body": f"{name} · {phone}"}},
            timeout=5,
        )
    except Exception as e:
        print(f"[FCM] {e}")

ONLINE_THRESHOLD_SECONDS = 90  # если heartbeat приходил в последние 90 сек — считаем "в сети"

def _is_online(m: "ManagerDB") -> bool:
    if not m.last_seen:
        return False
    return (datetime.utcnow() - m.last_seen).total_seconds() < ONLINE_THRESHOLD_SECONDS

def _tg_token() -> str:
    try:
        with SessionLocal() as db:
            s = db.get(SettingDB, "telegram_bot_token")
            if s and s.value:
                return s.value
    except Exception:
        pass
    return os.getenv("TELEGRAM_BOT_TOKEN", "")

def _tg_notify_chat_id() -> str:
    try:
        with SessionLocal() as db:
            s = db.get(SettingDB, "telegram_notify_chat_id")
            if s and s.value:
                return s.value
    except Exception:
        pass
    return os.getenv("TELEGRAM_CHAT_ID", "")

def _notify_new_lead(name: str, phone: str, source: Optional[str] = None, comment: Optional[str] = None):
    """Простое уведомление в Telegram-чат о новой заявке: имя, телефон, источник."""
    chat_id = _tg_notify_chat_id()
    if not chat_id or not _tg_token():
        return
    lines = [
        "🆕 *Новая заявка*",
        f"👤 Имя: {name or '—'}",
        f"📞 Телефон: {phone or '—'}",
        f"📍 Откуда: {source or '—'}",
    ]
    if comment:
        lines.append(f"💬 {comment}")
    _tg_send(chat_id, "\n".join(lines))

def _tg_send(chat_id: str, text: str, reply_markup: dict = None):
    tg_token = _tg_token()
    if not tg_token or not chat_id:
        return None
    try:
        payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
        if reply_markup:
            payload["reply_markup"] = reply_markup
        r = httpx.post(f"https://api.telegram.org/bot{tg_token}/sendMessage", json=payload, timeout=8)
        return r.json()
    except Exception as e:
        print(f"[TG send] {e}")
        return None

def _tg_edit(chat_id: str, message_id: int, text: str):
    tg_token = _tg_token()
    if not tg_token:
        return
    try:
        httpx.post(
            f"https://api.telegram.org/bot{tg_token}/editMessageText",
            json={"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "Markdown"},
            timeout=8,
        )
    except Exception as e:
        print(f"[TG edit] {e}")

def _tg_answer_callback(callback_id: str, text: str, show_alert: bool = False):
    tg_token = _tg_token()
    if not tg_token:
        return
    try:
        httpx.post(
            f"https://api.telegram.org/bot{tg_token}/answerCallbackQuery",
            json={"callback_query_id": callback_id, "text": text, "show_alert": show_alert},
            timeout=8,
        )
    except Exception as e:
        print(f"[TG cq] {e}")

def send_telegram_vyezd(lead: LeadDB, db: Session):
    """Уведомление всем подключённым клинерам при статусе 'Выезд на объект' —
    с кнопкой «Взять в работу» прямо в Telegram."""
    tg_token = _tg_token()
    if not tg_token:
        return
    cleaners = db.query(ManagerDB).filter(
        ManagerDB.role == "cleaner", ManagerDB.active == True,
        ManagerDB.telegram_chat_id.isnot(None), ManagerDB.telegram_chat_id != ""
    ).all()
    tg_chat_fallback = os.getenv("TELEGRAM_CHAT_ID", "")
    text = (
        f"🚗 *Новый выезд на объект!*\n\n"
        f"👤 Клиент: {lead.name}\n"
        f"📞 Телефон: {lead.phone}\n"
        f"💬 Комментарий: {lead.comment or '—'}\n\n"
        f"Нажмите кнопку ниже, чтобы взять заявку в работу 👇"
    )
    keyboard = {"inline_keyboard": [[{"text": "✅ Взять в работу", "callback_data": f"take:{lead.id}"}]]}
    if not cleaners:
        # Нет привязанных клинеров — шлём в общий чат (если задан) без кнопки
        if tg_chat_fallback:
            _tg_send(tg_chat_fallback, text)
        return
    for c in cleaners:
        res = _tg_send(c.telegram_chat_id, text, reply_markup=keyboard)
        if res and res.get("ok"):
            try:
                mid = res["result"]["message_id"]
                db.add(TelegramNotifyDB(lead_id=lead.id, chat_id=c.telegram_chat_id, message_id=mid))
            except Exception:
                pass
    db.commit()

def _broadcast_lead_taken(db: Session, lead: LeadDB):
    """Обновляет все разосланные клинерам сообщения — показывает, кто забрал заявку."""
    notifs = db.query(TelegramNotifyDB).filter(TelegramNotifyDB.lead_id == lead.id).all()
    if not notifs:
        return
    text = (
        f"✅ *Заявка взята в работу*\n\n"
        f"👤 Клиент: {lead.name}\n"
        f"📞 Телефон: {lead.phone}\n"
        f"🧹 Клинер: {lead.taken_by or '—'}"
    )
    for n in notifs:
        _tg_edit(n.chat_id, n.message_id, text)

def _log_activity(db, phone, event_type, payload=None, manager_id=None):
    try:
        db.add(ActivityDB(
            phone=phone, event_type=event_type,
            payload=json.dumps(payload, ensure_ascii=False) if payload else None,
            manager_id=manager_id
        ))
    except Exception:
        pass

# ─────────── EMAIL (IMAP) — заявки с почты становятся лидами ───────────

def _decode_mime(s):
    if not s:
        return ""
    try:
        parts = decode_header(s)
    except Exception:
        return s
    out = []
    for chunk, enc in parts:
        if isinstance(chunk, bytes):
            try:
                out.append(chunk.decode(enc or "utf-8", errors="ignore"))
            except Exception:
                out.append(chunk.decode("utf-8", errors="ignore"))
        else:
            out.append(chunk)
    return "".join(out)

def _get_email_body(msg) -> str:
    plain, html = None, None
    if msg.is_multipart():
        for part in msg.walk():
            disp = str(part.get("Content-Disposition") or "")
            if "attachment" in disp:
                continue
            ctype = part.get_content_type()
            try:
                payload = part.get_payload(decode=True)
            except Exception:
                payload = None
            if not payload:
                continue
            charset = part.get_content_charset() or "utf-8"
            try:
                text_val = payload.decode(charset, errors="ignore")
            except Exception:
                text_val = payload.decode("utf-8", errors="ignore")
            if ctype == "text/plain" and plain is None:
                plain = text_val
            elif ctype == "text/html" and html is None:
                html = text_val
    else:
        try:
            payload = msg.get_payload(decode=True)
            charset = msg.get_content_charset() or "utf-8"
            plain = payload.decode(charset, errors="ignore") if payload else (msg.get_payload() or "")
        except Exception:
            plain = msg.get_payload() if isinstance(msg.get_payload(), str) else ""
    if plain:
        return plain.strip()
    if html:
        return re.sub("<[^>]+>", " ", html).strip()
    return ""

def _imap_connect(acc: "EmailAccountDB"):
    if acc.imap_ssl:
        conn = imaplib.IMAP4_SSL(acc.imap_host, acc.imap_port)
    else:
        conn = imaplib.IMAP4(acc.imap_host, acc.imap_port)
    conn.login(acc.email, acc.password or "")
    return conn

# Отправители, письма от которых НЕ должны становиться заявками (лидами).
# Сюда попадают служебные/рекламные уведомления, а не реальные обращения клиентов.
# Проверка идёт по вхождению подстроки в адрес отправителя (регистр не важен).
IGNORED_SENDER_PATTERNS = [
    "hoster.by",         # уведомления хостинг-провайдера (тех. письма, счета и т.п.)
    "yandex-direct",     # Яндекс.Директ
    "yandex.direct",     # на случай другого написания домена/адреса
    "direct@yandex",     # рассылки/уведомления Яндекс.Директа
]

def _is_ignored_sender(from_addr: str) -> bool:
    """True, если письмо пришло от адреса из игнор-листа (см. IGNORED_SENDER_PATTERNS)."""
    if not from_addr:
        return False
    addr = from_addr.strip().lower()
    return any(pattern in addr for pattern in IGNORED_SENDER_PATTERNS)

def _email_account_test(acc: "EmailAccountDB"):
    try:
        conn = _imap_connect(acc)
        conn.select("INBOX", readonly=True)
        conn.logout()
        return True, None
    except Exception as e:
        return False, str(e)

def _poll_email_account(account_id: int) -> int:
    """Проверяет один ящик по IMAP и создаёт лиды по новым письмам. Возвращает число созданных лидов."""
    created = 0
    with SessionLocal() as db:
        acc = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
        if not acc or not acc.active or not acc.password:
            return 0
        conn = None
        try:
            conn = _imap_connect(acc)
            conn.select("INBOX")
            status, data = conn.uid("search", None, "ALL")
            if status != "OK":
                raise Exception("Ошибка IMAP SEARCH")
            uids = [int(u) for u in data[0].split()] if data and data[0] else []

            if acc.last_uid is None:
                # первая проверка — не создаём лиды из старых писем, просто фиксируем текущее состояние
                acc.last_uid = max(uids) if uids else 0
                acc.last_checked = datetime.utcnow()
                acc.last_error = None
                db.commit()
                return 0

            new_uids = sorted(u for u in uids if u > acc.last_uid)
            max_uid = acc.last_uid
            for uid in new_uids:
                try:
                    st, msg_data = conn.uid("fetch", str(uid), "(RFC822)")
                    if st != "OK" or not msg_data or not msg_data[0]:
                        max_uid = max(max_uid, uid)
                        continue
                    raw = msg_data[0][1]
                    msg = email_lib.message_from_bytes(raw)
                    subject = _decode_mime(msg.get("Subject")) or "(без темы)"
                    from_name, from_addr = parseaddr(msg.get("From", ""))
                    from_name = _decode_mime(from_name)
                    body = _get_email_body(msg)

                    # письма от самого ящика (авто-ответы и т.п.) не превращаем в лиды
                    if from_addr and acc.email and from_addr.strip().lower() == acc.email.strip().lower():
                        max_uid = max(max_uid, uid)
                        continue

                    # служебные/рекламные письма (hoster.by, yandex.direct и т.п.) — не заявки клиентов
                    if _is_ignored_sender(from_addr):
                        max_uid = max(max_uid, uid)
                        continue

                    phone_placeholder = f"email-{acc.id}-{uid}"
                    if not db.query(LeadDB).filter(LeadDB.phone == phone_placeholder).first():
                        pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == acc.pipeline_id).first()
                        stages = pipeline.stages.split("|") if pipeline else []
                        status_val = "Новая заявка" if "Новая заявка" in stages else (stages[0] if stages else "Новая заявка")
                        # Имя и email берём от отправителя письма, а не от ящика-получателя
                        lead_name = from_name or from_addr or acc.name or acc.email
                        lead_email = from_addr or None
                        lead = LeadDB(
                            name=lead_name,
                            phone=phone_placeholder,
                            status=status_val,
                            email=lead_email,
                            pipeline_id=acc.pipeline_id,
                            client_type=acc.client_type or "юр",
                            source="Email",
                            comment=f"Тема письма: {subject}",
                        )
                        db.add(lead)
                        db.flush()
                        note_text = (
                            f"📧 Письмо на {acc.email}\n"
                            f"От: {from_name or '—'} <{from_addr or '—'}>\n"
                            f"Тема: {subject}\n\n{body[:5000]}"
                        )
                        db.add(NoteDB(phone=phone_placeholder, text=note_text, note_type="email"))
                        _log_activity(db, phone_placeholder, "lead_created",
                                      {"name": lead.name, "source": "Email", "subject": subject})
                        db.commit()
                        _notify_new_lead(lead.name, phone_placeholder, "Email")
                        created += 1
                    max_uid = max(max_uid, uid)
                except Exception as e:
                    print(f"[email] ошибка обработки письма uid={uid} ({acc.email}): {e}")
                    max_uid = max(max_uid, uid)
                    continue

            acc.last_uid = max_uid
            acc.last_checked = datetime.utcnow()
            acc.last_error = None
            acc.leads_created = (acc.leads_created or 0) + created
            db.commit()
        except Exception as e:
            acc.last_error = str(e)
            acc.last_checked = datetime.utcnow()
            db.commit()
            print(f"[email] ошибка подключения к {acc.email}: {e}")
        finally:
            if conn is not None:
                try:
                    conn.logout()
                except Exception:
                    pass
    return created

def _idle_watch_account(account_id: int):
    """
    Фоновый поток для одного ящика.
    Использует IMAP IDLE (RFC 2177): сервер сам присылает EXISTS/RECENT
    при новом письме — реагируем мгновенно, без периодического опроса.
    При любой ошибке ждём IDLE_RECONNECT_SECONDS и переподключаемся.
    """
    IDLE_TIMEOUT   = int(os.getenv("IMAP_IDLE_TIMEOUT",   "1200"))  # 20 мин — обновляем IDLE до истечения 30 мин
    RECONNECT_WAIT = int(os.getenv("IDLE_RECONNECT_SECONDS", "30"))

    while True:
        conn = None
        try:
            with SessionLocal() as db:
                acc = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
                if not acc or not acc.active or not acc.password:
                    time.sleep(RECONNECT_WAIT)
                    continue

            conn = _imap_connect(acc)
            conn.select("INBOX")

            # ── baseline при первом подключении ──────────────────────────
            with SessionLocal() as db:
                acc = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
                if acc.last_uid is None:
                    status, data = conn.uid("search", None, "ALL")
                    uids = [int(u) for u in data[0].split()] if data and data[0] else []
                    acc.last_uid = max(uids) if uids else 0
                    acc.last_checked = datetime.utcnow()
                    acc.last_error = None
                    db.commit()
                    print(f"[email-idle] baseline {acc.email}: last_uid={acc.last_uid}")

            print(f"[email-idle] IDLE запущен для {acc.email}")

            idle_start = time.monotonic()
            # Отправляем команду IDLE
            tag = conn._new_tag()
            conn.send(tag + b" IDLE\r\n")
            resp = conn.readline()                       # ожидаем "+ idling"
            if not resp.startswith(b"+"):
                raise Exception(f"Сервер не поддерживает IDLE: {resp!r}")

            while True:
                elapsed = time.monotonic() - idle_start
                remaining = IDLE_TIMEOUT - elapsed
                if remaining <= 0:
                    # Обновляем IDLE — отправляем DONE и снова IDLE
                    conn.send(b"DONE\r\n")
                    # читаем ответ OK на предыдущий IDLE
                    while True:
                        line = conn.readline()
                        if tag in line or b"OK" in line or b"BAD" in line or b"NO" in line:
                            break
                    tag = conn._new_tag()
                    conn.send(tag + b" IDLE\r\n")
                    conn.readline()                     # "+ idling"
                    idle_start = time.monotonic()
                    continue

                # Ждём данных от сервера (неблокирующий select)
                import select as _select
                rlist, _, _ = _select.select([conn.socket()], [], [], min(remaining, 60))
                if not rlist:
                    continue                            # таймаут select, продолжаем ждать

                # Читаем строку-событие
                line = conn.readline()
                if not line:
                    raise Exception("Сервер закрыл соединение")

                # EXISTS или RECENT = пришло новое письмо
                if b"EXISTS" in line or b"RECENT" in line:
                    print(f"[email-idle] {acc.email} — новое письмо ({line.strip()!r}), обрабатываем")
                    # Выходим из IDLE
                    conn.send(b"DONE\r\n")
                    while True:
                        done_line = conn.readline()
                        if tag in done_line or b"OK" in done_line or b"BAD" in done_line:
                            break

                    # Обрабатываем новые письма
                    created = _poll_email_account(account_id)
                    if created:
                        print(f"[email-idle] {acc.email} — создано {created} лид(ов)")

                    # Возвращаемся в IDLE
                    tag = conn._new_tag()
                    conn.send(tag + b" IDLE\r\n")
                    conn.readline()
                    idle_start = time.monotonic()

        except Exception as e:
            print(f"[email-idle] ошибка ({acc.email if 'acc' in dir() else account_id}): {e} — reconnect через {RECONNECT_WAIT}s")
            try:
                with SessionLocal() as db:
                    a = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
                    if a:
                        a.last_error = str(e)
                        a.last_checked = datetime.utcnow()
                        db.commit()
            except Exception:
                pass
            time.sleep(RECONNECT_WAIT)
        finally:
            if conn is not None:
                try:
                    conn.logout()
                except Exception:
                    pass


# Реестр запущенных потоков IDLE: account_id -> Thread
_idle_threads: dict = {}
_idle_lock = threading.Lock()


def _ensure_idle_threads():
    """
    Запускает IDLE-поток для каждого активного ящика, которого ещё нет в реестре.
    Вызывается при старте и при изменении списка аккаунтов.
    """
    with SessionLocal() as db:
        accounts = db.query(EmailAccountDB).filter(EmailAccountDB.active == True).all()
        ids = [a.id for a in accounts]

    with _idle_lock:
        for aid in ids:
            if aid not in _idle_threads or not _idle_threads[aid].is_alive():
                t = threading.Thread(target=_idle_watch_account, args=(aid,), daemon=True,
                                     name=f"imap-idle-{aid}")
                t.start()
                _idle_threads[aid] = t
                print(f"[email-idle] поток запущен для account_id={aid}")


def _email_poll_loop():
    """
    Оставлено как fallback для серверов без IDLE.
    Также служит watchdog — каждые EMAIL_POLL_SECONDS перезапускает упавшие потоки.
    """
    interval = int(os.getenv("EMAIL_POLL_SECONDS", "60"))
    while True:
        try:
            _ensure_idle_threads()
        except Exception as e:
            print(f"[email] watchdog ошибка: {e}")
        time.sleep(interval)

def _email_account_to_dict(a: "EmailAccountDB"):
    return {
        "id": a.id, "name": a.name, "email": a.email,
        "has_password": bool(a.password),
        "imap_host": a.imap_host, "imap_port": a.imap_port, "imap_ssl": a.imap_ssl,
        "smtp_host": a.smtp_host, "smtp_port": a.smtp_port, "smtp_ssl": a.smtp_ssl,
        "pipeline_id": a.pipeline_id, "client_type": a.client_type,
        "active": a.active,
        "last_checked": a.last_checked.isoformat() if a.last_checked else None,
        "last_error": a.last_error,
        "leads_created": a.leads_created or 0,
        "baseline_done": a.last_uid is not None,
    }

# ─────────── ВНУТРЕННЯЯ РАССЫЛКА (SMTP) ───────────

MAIL_SEND_INTERVAL = float(os.getenv("MAIL_SEND_INTERVAL_SECONDS", "2"))   # пауза между письмами, антиспам-троттлинг

def _render_template(text: str, name: str) -> str:
    return (text or "").replace("{{name}}", name or "").replace("{{имя}}", name or "")

def _send_smtp_mail(account: "EmailAccountDB", to_email: str, subject: str, body: str):
    """Отправляет одно письмо через SMTP-настройки корпоративного ящика (те же,
    что используются для IMAP-приёма — секция «Почта» в Интеграциях)."""
    if not account.smtp_host or not account.password:
        raise RuntimeError("Не настроен SMTP-хост или пароль для этого ящика")
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = account.email
    msg["To"] = to_email
    msg.attach(MIMEText(body, "plain", "utf-8"))
    port = account.smtp_port or (465 if account.smtp_ssl else 587)
    if account.smtp_ssl:
        with smtplib.SMTP_SSL(account.smtp_host, port, timeout=15) as s:
            s.login(account.email, account.password)
            s.sendmail(account.email, [to_email], msg.as_string())
    else:
        with smtplib.SMTP(account.smtp_host, port, timeout=15) as s:
            s.starttls()
            s.login(account.email, account.password)
            s.sendmail(account.email, [to_email], msg.as_string())

def _send_campaign_sync(db: Session, campaign_id: str) -> dict:
    """Отправляет все письма рассылки прямо здесь и сейчас, синхронно, одно за
    другим, через SMTP — без очереди, без фонового потока. Вызывается прямо
    из обработчика POST /mailings, поэтому HTTP-ответ приходит только когда
    все письма уже реально отправлены (или провалены)."""
    items = (db.query(MailQueueDB)
               .filter(MailQueueDB.campaign_id == campaign_id, MailQueueDB.status == "pending")
               .order_by(MailQueueDB.id.asc())
               .all())
    sent, failed = 0, 0
    for idx, item in enumerate(items):
        account = db.query(EmailAccountDB).filter(EmailAccountDB.id == item.email_account_id).first()
        item.attempts = 1
        try:
            if not account:
                raise RuntimeError("Почтовый ящик отправителя удалён")
            _send_smtp_mail(account, item.to_email, item.subject, item.body)
            item.status = "sent"
            item.sent_at = datetime.utcnow()
            item.error = None
            sent += 1
        except Exception as e:
            item.status = "failed"
            item.error = str(e)
            failed += 1
        db.commit()
        if idx < len(items) - 1:
            time.sleep(MAIL_SEND_INTERVAL)   # троттлинг, чтобы не словить блокировку от провайдера
    return {"sent": sent, "failed": failed}

def _lead_to_dict(lead: LeadDB, db: Session):
    tags = db.query(TagDB).join(LeadTagDB, LeadTagDB.tag_id == TagDB.tag_id).filter(LeadTagDB.lead_id == lead.id).all()
    custom_rows = db.query(LeadCustomValueDB, CustomFieldDB).join(
        CustomFieldDB, CustomFieldDB.field_id == LeadCustomValueDB.field_id
    ).filter(LeadCustomValueDB.lead_id == lead.id).all()
    custom = {f.name: v.value for v, f in custom_rows}
    return {
        "id":             lead.id,
        "name":           lead.name,
        "phone":          lead.phone,
        "status":         lead.status,
        "loss_reason_id": lead.loss_reason_id,
        "telegram":       lead.telegram,
        "viber":          lead.viber,
        "vk":             lead.vk,
        "whatsapp":       lead.whatsapp,
        "email":          lead.email,
        "pipeline_id":    lead.pipeline_id,
        "manager_id":     lead.manager_id,
        "client_type":    lead.client_type,
        "source":         lead.source,
        "company_name":   lead.company_name,
        "address":        lead.address,
        "budget":         lead.budget,
        "comment":        lead.comment,
        "taken_by":       lead.taken_by,
        "taken_by_manager_id": lead.taken_by_manager_id,
        "tags":           [{"tag_id": t.tag_id, "name": t.name, "color": t.color} for t in tags],
        "custom":         custom,
    }

# ─────────── LEADS ───────────

@app.post("/leads")
def create_lead(lead: Lead, db: Session = Depends(get_db)):
    if db.query(LeadDB).filter(LeadDB.phone == lead.phone).first():
        return {"error": "Лид с таким номером телефона уже существует"}
    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == lead.pipeline_id).first()
    if not pipeline:
        return {"error": f"Воронка с id {lead.pipeline_id} не найдена"}
    if lead.manager_id is not None:
        if not db.query(ManagerDB).filter(ManagerDB.manager_id == lead.manager_id).first():
            return {"error": f"Менеджер с id {lead.manager_id} не найден"}
    stages = pipeline.stages.split("|")
    status = lead.status if (lead.status and lead.status in stages) else stages[0]

    db_lead = LeadDB(
        name=lead.name, phone=lead.phone, status=status,
        loss_reason_id=lead.loss_reason_id,
        telegram=_clean(lead.telegram), viber=_clean(lead.viber), vk=_clean(lead.vk),
        whatsapp=_clean(lead.whatsapp), email=_clean(lead.email),
        pipeline_id=lead.pipeline_id, manager_id=lead.manager_id,
        client_type=lead.client_type, source=_clean(lead.source),
        company_name=_clean(lead.company_name), address=_clean(lead.address), budget=_clean(lead.budget),
        comment=_clean(lead.comment),
    )
    db.add(db_lead)
    db.flush()
    for tag_id in (lead.tag_ids or []):
        if db.query(TagDB).filter(TagDB.tag_id == tag_id).first():
            db.add(LeadTagDB(lead_id=db_lead.id, tag_id=tag_id))
    db.commit()
    db.refresh(db_lead)
    _log_activity(db, lead.phone, "lead_created", {"name": lead.name, "source": lead.source}, lead.manager_id)
    db.commit()
    _notify_new_lead(lead.name, lead.phone, lead.source)
    return {"message": "Лид успешно добавлен", "data": _lead_to_dict(db_lead, db)}

# ─────────── LEADS LIST — с фильтрами/сортировкой/пагинацией ───────────

@app.get("/leads")
def list_leads(
    db: Session = Depends(get_db),
    search:        Optional[str] = None,
    pipeline_id:   Optional[int] = None,
    manager_id:    Optional[int] = None,
    status:        Optional[str] = None,
    source:        Optional[str] = None,
    tag_id:        Optional[int] = None,
    client_type:   Optional[str] = None,
    created_from:  Optional[int] = None,
    created_to:    Optional[int] = None,
    sort_by:       Optional[str] = "created",
    order:         Optional[str] = "desc",
    page:          int = 1,
    per_page:      int = 50,
):
    q = db.query(LeadDB)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(LeadDB.name.ilike(like), LeadDB.phone.ilike(like),
                         LeadDB.email.ilike(like), LeadDB.company_name.ilike(like)))
    if pipeline_id:  q = q.filter(LeadDB.pipeline_id == pipeline_id)
    if manager_id:   q = q.filter(LeadDB.manager_id == manager_id)
    if status:       q = q.filter(LeadDB.status == status)
    if source:       q = q.filter(LeadDB.source == source)
    if client_type:  q = q.filter(LeadDB.client_type == client_type)
    if created_from: q = q.filter(LeadDB.id >= created_from)  # surrogate for "after this id" on sqlite
    if created_to:   q = q.filter(LeadDB.id <= created_to)
    if tag_id:
        lead_ids = [r.lead_id for r in db.query(LeadTagDB.lead_id).filter(LeadTagDB.tag_id == tag_id).all()]
        q = q.filter(LeadDB.id.in_(lead_ids))

    sort_map = {
        "created": LeadDB.id,
        "name": LeadDB.name,
        "status": LeadDB.status,
    }
    col = sort_map.get(sort_by, LeadDB.id)
    q = q.order_by(col.desc() if order == "desc" else col.asc())

    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()
    return {"items": [_lead_to_dict(l, db) for l in items], "total": total, "page": page, "per_page": per_page}

@app.put("/leads/update-status")
def update_status(phone: str, new_status: str, loss_reason_id: Optional[int] = None,
                  manager_id: Optional[int] = None, db: Session = Depends(get_db)):
    lead = db.query(LeadDB).filter(LeadDB.phone == phone).first()
    if not lead:
        return {"error": "Клиент с таким номером не найден."}

    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == lead.pipeline_id).first()
    allowed = pipeline.stages.split("|") if pipeline else ["Новый клиент", "В работе", "Отказ"]
    if new_status not in allowed:
        return {"error": f"Статус '{new_status}' недопустим"}
    if new_status == "Отказ" and loss_reason_id is None:
        return {"error": "Укажите loss_reason_id для отказа"}
    if loss_reason_id is not None and not db.query(LossReasonDB).filter(LossReasonDB.reason_id == loss_reason_id).first():
        return {"error": "Причина отказа не найдена"}

    old = lead.status
    lead.status = new_status
    lead.loss_reason_id = loss_reason_id
    db.add(StageHistoryDB(phone=phone, old_status=old, new_status=new_status, manager_id=manager_id))
    _log_activity(db, phone, "stage_change", {"from": old, "to": new_status, "loss_reason_id": loss_reason_id}, manager_id)
    db.commit()
    # Telegram-уведомление при переходе на "Выезд на объект"
    if new_status == "Выезд на объект" and old != "Выезд на объект":
        send_telegram_vyezd(lead, db)
    return {"message": "Статус изменён", "old": old, "new": new_status}

# ─────────── TAKE LEAD (клинер берёт в работу) ───────────

class TakeWork(BaseModel):
    cleaner_name: Optional[str] = None
    manager_id:   Optional[int] = None

@app.post("/leads/{lead_id}/take")
def take_lead(lead_id: int, body: TakeWork, db: Session = Depends(get_db)):
    lead = db.query(LeadDB).filter(LeadDB.id == lead_id).first()
    if not lead:
        return JSONResponse(status_code=404, content={"error": "Лид не найден"})
    if lead.status != "Выезд на объект":
        return JSONResponse(status_code=400, content={"error": "Лид не в статусе 'Выезд на объект'"})
    if lead.taken_by:
        return JSONResponse(status_code=400, content={"error": f"Уже взято: {lead.taken_by}"})

    cleaner_name = None
    manager_id = None
    if body.manager_id is not None:
        mgr = db.query(ManagerDB).filter(ManagerDB.manager_id == body.manager_id).first()
        if not mgr:
            return JSONResponse(status_code=404, content={"error": "Клинер не найден"})
        cleaner_name = mgr.name
        manager_id = mgr.manager_id
    elif body.cleaner_name:
        cleaner_name = body.cleaner_name.strip()
    if not cleaner_name:
        return JSONResponse(status_code=422, content={"error": "Не указан клинер"})

    old = lead.status
    lead.status = "Взято в работу"
    lead.taken_by = cleaner_name
    lead.taken_by_manager_id = manager_id
    db.add(StageHistoryDB(phone=lead.phone, old_status=old, new_status="Взято в работу", manager_id=manager_id))
    _log_activity(db, lead.phone, "stage_change", {"from": old, "to": "Взято в работу", "cleaner": cleaner_name}, manager_id)
    db.commit()
    db.refresh(lead)
    _broadcast_lead_taken(db, lead)
    return {"message": "Взято в работу", "data": _lead_to_dict(lead, db)}

@app.put("/leads/update")
def update_lead(phone: str, payload: LeadUpdate, db: Session = Depends(get_db)):
    """amo-style редактирование любых полей лида одной кнопкой."""
    lead = db.query(LeadDB).filter(LeadDB.phone == phone).first()
    if not lead:
        return {"error": "Не найден"}
    changed = []
    if payload.name is not None and payload.name != lead.name:
        changed.append(("name", lead.name, payload.name)); lead.name = payload.name
    for fld in ("telegram","viber","vk","whatsapp","email","source","client_type","company_name","address","budget","comment"):
        v = getattr(payload, fld)
        if v is not None and getattr(lead, fld) != v:
            changed.append((fld, getattr(lead, fld), v))
            setattr(lead, fld, _clean(v))
    if payload.pipeline_id is not None and payload.pipeline_id != lead.pipeline_id:
        changed.append(("pipeline_id", lead.pipeline_id, payload.pipeline_id))
        lead.pipeline_id = payload.pipeline_id
        pl = db.query(PipelineDB).filter(PipelineDB.pipeline_id == payload.pipeline_id).first()
        if pl: lead.status = pl.stages.split("|")[0]
    if payload.manager_id is not None and payload.manager_id != lead.manager_id:
        changed.append(("manager_id", lead.manager_id, payload.manager_id))
        lead.manager_id = payload.manager_id
    if payload.tag_ids is not None:
        current = {r.tag_id for r in db.query(LeadTagDB).filter(LeadTagDB.lead_id == lead.id).all()}
        new = set(payload.tag_ids)
        for t in current - new:
            db.query(LeadTagDB).filter(LeadTagDB.lead_id == lead.id, LeadTagDB.tag_id == t).delete()
        for t in new - current:
            if db.query(TagDB).filter(TagDB.tag_id == t).first():
                db.add(LeadTagDB(lead_id=lead.id, tag_id=t))
        changed.append(("tags", list(current), list(new)))
    # custom fields
    if payload.custom:
        for fname, val in payload.custom.items():
            fld = db.query(CustomFieldDB).filter(CustomFieldDB.name == fname).first()
            if not fld: continue
            existing = db.query(LeadCustomValueDB).filter(
                LeadCustomValueDB.lead_id == lead.id,
                LeadCustomValueDB.field_id == fld.field_id
            ).first()
            if existing:
                existing.value = val
            else:
                db.add(LeadCustomValueDB(lead_id=lead.id, field_id=fld.field_id, value=val))

    _log_activity(db, phone, "lead_updated", {"changes": changed})
    db.commit()
    return {"message": "Сохранено", "data": _lead_to_dict(lead, db)}

@app.delete("/leads/delete")
def delete_lead(phone: str, db: Session = Depends(get_db)):
    lead = db.query(LeadDB).filter(LeadDB.phone == phone).first()
    if not lead:
        return {"error": "Телефона нет в базе"}
    db.query(LeadTagDB).filter(LeadTagDB.lead_id == lead.id).delete()
    db.query(LeadCustomValueDB).filter(LeadCustomValueDB.lead_id == lead.id).delete()
    db.query(NoteDB).filter(NoteDB.phone == phone).delete()
    db.query(TaskDB).filter(TaskDB.phone == phone).delete()
    db.query(ActivityDB).filter(ActivityDB.phone == phone).delete()
    db.query(StageHistoryDB).filter(StageHistoryDB.phone == phone).delete()
    db.delete(lead)
    db.commit()
    return {"message": "Клиент удалён"}

# ─────────── BULK ACTIONS ───────────

@app.post("/leads/bulk")
def bulk_action(payload: BulkUpdate, db: Session = Depends(get_db)):
    if not payload.phones:
        return {"error": "Не выбраны лиды"}
    leads = db.query(LeadDB).filter(LeadDB.phone.in_(payload.phones)).all()
    if not leads:
        return {"error": "Ничего не найдено"}

    if payload.delete:
        for l in leads:
            db.query(LeadTagDB).filter(LeadTagDB.lead_id == l.id).delete()
            db.query(LeadCustomValueDB).filter(LeadCustomValueDB.lead_id == l.id).delete()
            db.delete(l)
        db.commit()
        return {"message": f"Удалено: {len(leads)}"}

    for l in leads:
        if payload.set_pipeline is not None:
            l.pipeline_id = payload.set_pipeline
            pl = db.query(PipelineDB).filter(PipelineDB.pipeline_id == payload.set_pipeline).first()
            if pl: l.status = pl.stages.split("|")[0]
        if payload.set_manager is not None:
            l.manager_id = payload.set_manager
        if payload.set_status is not None:
            old = l.status; l.status = payload.set_status
            db.add(StageHistoryDB(phone=l.phone, old_status=old, new_status=payload.set_status))
        if payload.add_tags:
            current = {r.tag_id for r in db.query(LeadTagDB).filter(LeadTagDB.lead_id == l.id).all()}
            for t in payload.add_tags:
                if t not in current:
                    db.add(LeadTagDB(lead_id=l.id, tag_id=t))
        if payload.remove_tags:
            db.query(LeadTagDB).filter(LeadTagDB.lead_id == l.id, LeadTagDB.tag_id.in_(payload.remove_tags)).delete(synchronize_session=False)
        _log_activity(db, l.phone, "bulk_update", {"by": payload.set_status or "n/a"})

    db.commit()
    return {"message": f"Обновлено: {len(leads)}", "count": len(leads)}

# ─────────── NOTES ───────────

@app.post("/notes")
def add_note(note: Note, manager_id: Optional[int] = None, db: Session = Depends(get_db)):
    if not db.query(LeadDB).filter(LeadDB.phone == note.phone).first():
        return {"error": "Лид не найден"}
    n = NoteDB(phone=note.phone, text=note.text, note_type=note.note_type, manager_id=manager_id)
    db.add(n)
    _log_activity(db, note.phone, "note_added", {"type": note.note_type, "len": len(note.text)}, manager_id)
    db.commit()
    return {"message": "Заметка добавлена", "data": {"note_id": n.id, "created_at": n.created_at.isoformat()}}

@app.get("/notes")
def list_notes(phone: str, db: Session = Depends(get_db)):
    return db.query(NoteDB).filter(NoteDB.phone == phone).order_by(NoteDB.created_at.desc()).all()

@app.delete("/notes/{note_id}")
def delete_note(note_id: int, db: Session = Depends(get_db)):
    n = db.query(NoteDB).filter(NoteDB.id == note_id).first()
    if not n:
        return {"error": "Не найдено"}
    db.delete(n); db.commit()
    return {"message": "Удалено"}

# ─────────── TAGS ───────────

@app.get("/tags")
def list_tags(db: Session = Depends(get_db)):
    return [{"tag_id": t.tag_id, "name": t.name, "color": t.color} for t in db.query(TagDB).order_by(TagDB.name).all()]

@app.post("/tags")
def create_tag(tag: Tag, db: Session = Depends(get_db)):
    if db.query(TagDB).filter(TagDB.name == tag.name).first():
        return {"error": "Тег с таким именем уже существует"}
    t = TagDB(name=tag.name, color=tag.color)
    db.add(t); db.commit()
    return {"message": "Тег создан", "data": {"tag_id": t.tag_id, "name": t.name, "color": t.color}}

@app.delete("/tags/{tag_id}")
def delete_tag(tag_id: int, db: Session = Depends(get_db)):
    t = db.query(TagDB).filter(TagDB.tag_id == tag_id).first()
    if not t: return {"error": "Не найден"}
    db.query(LeadTagDB).filter(LeadTagDB.tag_id == tag_id).delete()
    db.delete(t); db.commit()
    return {"message": "Тег удалён"}

# ─────────── CUSTOM FIELDS ───────────

@app.get("/custom-fields")
def list_custom_fields(pipeline_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(CustomFieldDB)
    if pipeline_id:
        q = q.filter(or_(CustomFieldDB.pipeline_id == pipeline_id, CustomFieldDB.pipeline_id.is_(None)))
    fields = q.order_by(CustomFieldDB.sort).all()
    out = []
    for f in fields:
        out.append({
            "field_id":   f.field_id,
            "pipeline_id":f.pipeline_id,
            "name":       f.name,
            "field_type": f.field_type,
            "options":    f.options.split("|") if f.options else None,
            "required":   f.required,
            "sort":       f.sort,
        })
    return out

@app.post("/custom-fields")
def create_custom_field(field: CustomField, db: Session = Depends(get_db)):
    cf = CustomFieldDB(
        pipeline_id=field.pipeline_id, name=field.name, field_type=field.field_type,
        options=field.options, required=field.required, sort=field.sort
    )
    db.add(cf); db.commit()
    return {"message": "Поле создано", "data": {"field_id": cf.field_id}}

@app.delete("/custom-fields/{field_id}")
def delete_custom_field(field_id: int, db: Session = Depends(get_db)):
    cf = db.query(CustomFieldDB).filter(CustomFieldDB.field_id == field_id).first()
    if not cf: return {"error": "Не найдено"}
    db.query(LeadCustomValueDB).filter(LeadCustomValueDB.field_id == field_id).delete()
    db.delete(cf); db.commit()
    return {"message": "Поле удалено"}

# ─────────── HISTORY / ACTIVITY ───────────

@app.get("/pipeline/history")
def history(db: Session = Depends(get_db), phone: Optional[str] = None, limit: int = 500):
    q = db.query(StageHistoryDB)
    if phone: q = q.filter(StageHistoryDB.phone == phone)
    return q.order_by(StageHistoryDB.changed_at.desc()).limit(limit).all()

@app.get("/activity")
def activity(phone: str, db: Session = Depends(get_db)):
    rows = db.query(ActivityDB).filter(ActivityDB.phone == phone).order_by(ActivityDB.created_at.desc()).all()
    return [{"event_type": r.event_type, "payload": r.payload, "manager_id": r.manager_id, "created_at": r.created_at.isoformat()} for r in rows]

# ─────────── ANALYTICS ───────────

@app.get("/analytics/pipeline-conversion")
def pipeline_conversion(db: Session = Depends(get_db)):
    pipelines = {p.pipeline_id: p.stages.split("|") for p in db.query(PipelineDB).all()}
    leads = db.query(LeadDB).all()
    total = len(leads)
    if total == 0:
        return {"conversion_rate": 0.0, "total": 0, "successful": 0, "by_pipeline": []}
    successful = sum(
        1 for l in leads
        if (st := pipelines.get(l.pipeline_id))
        and l.status == (st[-2] if st[-1] == "Отказ" else st[-1])
    )
    by_pipeline = []
    for pid, st in pipelines.items():
        in_pl = [l for l in leads if l.pipeline_id == pid]
        succ = sum(1 for l in in_pl if l.status == (st[-2] if st[-1] == "Отказ" else st[-1]))
        refused = sum(1 for l in in_pl if l.status == "Отказ")
        by_pipeline.append({
            "pipeline_id": pid,
            "total": len(in_pl),
            "successful": succ,
            "refused": refused,
            "conversion": round((succ / len(in_pl) * 100) if in_pl else 0, 2),
        })
    return {
        "total": total, "successful": successful,
        "conversion_rate": round(successful / total * 100, 2),
        "by_pipeline": by_pipeline,
    }

@app.get("/analytics/funnel")
def analytics_funnel(pipeline_id: int, db: Session = Depends(get_db)):
    """Воронка конверсии по этапам — классическая amoCRM-аналитика."""
    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == pipeline_id).first()
    if not pipeline: return {"error": "Воронка не найдена"}
    stages = pipeline.stages.split("|")
    leads = db.query(LeadDB).filter(LeadDB.pipeline_id == pipeline_id).all()
    rows = []
    for i, st in enumerate(stages):
        if st == "Отказ":
            count = sum(1 for l in leads if l.status == "Отказ")
        else:
            # лид "дошёл до этого этапа" если его статус >= этот этап по порядку
            idx = stages.index(st)
            count = 0
            for l in leads:
                if l.status == "Отказ":
                    continue
                if l.status in stages:
                    if stages.index(l.status) >= idx:
                        count += 1
        rows.append({"stage": st, "count": count})
    return {"pipeline_id": pipeline_id, "pipeline_name": pipeline.name, "stages": rows}

@app.get("/analytics/by-manager")
def analytics_by_manager(db: Session = Depends(get_db)):
    managers = db.query(ManagerDB).filter(ManagerDB.active == True).all()
    result = []
    for m in managers:
        leads = db.query(LeadDB).filter(LeadDB.manager_id == m.manager_id).all()
        pipelines = {p.pipeline_id: p.stages.split("|") for p in db.query(PipelineDB).all()}
        succ = sum(1 for l in leads if (st := pipelines.get(l.pipeline_id)) and l.status == (st[-2] if st[-1] == "Отказ" else st[-1]))
        result.append({
            "manager_id": m.manager_id, "name": m.name, "role": m.role,
            "total": len(leads), "successful": succ,
            "conversion": round((succ / len(leads) * 100) if leads else 0, 2),
            "active_tasks": db.query(TaskDB).filter(TaskDB.manager_id == m.manager_id, TaskDB.is_completed == False).count(),
        })
    return result

@app.get("/analytics/by-source")
def analytics_by_source(db: Session = Depends(get_db)):
    pipelines = {p.pipeline_id: p.stages.split("|") for p in db.query(PipelineDB).all()}
    rows = db.query(LeadDB.source, func.count(LeadDB.id)).group_by(LeadDB.source).all()
    out = []
    for source, total in rows:
        leads = db.query(LeadDB).filter(LeadDB.source == source).all() if source else []
        succ = sum(1 for l in leads if (st := pipelines.get(l.pipeline_id)) and l.status == (st[-2] if st[-1] == "Отказ" else st[-1]))
        out.append({"source": source or "Не указан", "total": total, "successful": succ,
                    "conversion": round((succ / total * 100) if total else 0, 2)})
    out.sort(key=lambda x: -x["total"])
    return out

@app.get("/analytics/dashboard")
def analytics_dashboard(db: Session = Depends(get_db)):
    """Сводка цифр для главного экрана — всё одним запросом."""
    leads = db.query(LeadDB).all()
    pipelines = {p.pipeline_id: p.stages.split("|") for p in db.query(PipelineDB).all()}
    tasks_pending = db.query(TaskDB).filter(TaskDB.is_completed == False).count()
    overdue = db.query(TaskDB).filter(TaskDB.is_completed == False, TaskDB.deadline < int(datetime.utcnow().timestamp())).count()
    active = sum(1 for l in leads if l.status != "Отказ")
    succ = sum(1 for l in leads if (st := pipelines.get(l.pipeline_id)) and l.status == (st[-2] if st[-1] == "Отказ" else st[-1]))
    refused = sum(1 for l in leads if l.status == "Отказ")
    # новые за сегодня (по id считаем грубо как последние 50 id — простой proxy)
    new_today = db.query(LeadDB).filter(LeadDB.id >= max(1, (max((l.id for l in leads), default=0)) - 20)).count() if leads else 0
    return {
        "total": len(leads), "active": active, "successful": succ, "refused": refused,
        "conversion": round((succ / len(leads) * 100) if leads else 0, 2),
        "tasks_pending": tasks_pending, "tasks_overdue": overdue,
        "new_recent": new_today,
    }

# ─────────── TASKS ───────────

@app.post("/tasks")
def create_task(task: Task, db: Session = Depends(get_db)):
    if not db.query(LeadDB).filter(LeadDB.phone == task.phone).first():
        return {"error": "Лид не найден"}
    if db.query(TaskDB).filter(TaskDB.task_id == task.task_id).first():
        return {"error": "task_id уже занят"}
    t = TaskDB(task_id=task.task_id, phone=task.phone, description=task.description,
               deadline=task.deadline, is_completed=task.is_completed, task_type=task.task_type)
    db.add(t); db.commit()
    _log_activity(db, task.phone, "task_added", {"task_id": task.task_id, "type": task.task_type})
    db.commit()
    return {"message": "Задача создана", "data": {"task_id": t.task_id}}

@app.get("/tasks/all-pending")
def tasks_pending(db: Session = Depends(get_db), manager_id: Optional[int] = None, phone: Optional[str] = None):
    q = db.query(TaskDB).filter(TaskDB.is_completed == False)
    if manager_id: q = q.filter(TaskDB.manager_id == manager_id)
    if phone: q = q.filter(TaskDB.phone == phone)
    return q.order_by(TaskDB.deadline.asc()).all()

@app.get("/tasks/all")
def tasks_all(db: Session = Depends(get_db), phone: Optional[str] = None, limit: int = 500):
    q = db.query(TaskDB)
    if phone: q = q.filter(TaskDB.phone == phone)
    return q.order_by(TaskDB.deadline.desc()).limit(limit).all()

@app.put("/tasks/complete")
def complete_task(task_id: int, db: Session = Depends(get_db)):
    t = db.query(TaskDB).filter(TaskDB.task_id == task_id).first()
    if not t: return {"error": "Не найдено"}
    t.is_completed = True
    _log_activity(db, t.phone, "task_completed", {"task_id": task_id})
    db.commit()
    return {"message": "Готово"}

@app.put("/tasks/update")
def update_task(payload: TaskUpdate, task_id: int, db: Session = Depends(get_db)):
    t = db.query(TaskDB).filter(TaskDB.task_id == task_id).first()
    if not t: return {"error": "Не найдено"}
    if payload.description is not None: t.description = payload.description
    if payload.deadline is not None: t.deadline = payload.deadline
    if payload.is_completed is not None: t.is_completed = payload.is_completed
    db.commit()
    return {"message": "Сохранено"}

@app.delete("/tasks/{task_id}")
def delete_task(task_id: int, db: Session = Depends(get_db)):
    t = db.query(TaskDB).filter(TaskDB.task_id == task_id).first()
    if not t: return {"error": "Не найдено"}
    db.delete(t); db.commit()
    return {"message": "Удалено"}

# ─────────── LOSS REASONS ───────────

@app.post("/loss-reasons")
def create_loss_reason(loss: LossReason, db: Session = Depends(get_db)):
    if db.query(LossReasonDB).filter(LossReasonDB.reason_id == loss.reason_id).first():
        return {"error": "ID занят"}
    db.add(LossReasonDB(reason_id=loss.reason_id, title=loss.title)); db.commit()
    return {"message": "Создано"}

@app.get("/loss-reasons")
def list_loss_reasons(db: Session = Depends(get_db)):
    return db.query(LossReasonDB).order_by(LossReasonDB.reason_id).all()

@app.delete("/loss-reasons/{reason_id}")
def delete_loss_reason(reason_id: int, db: Session = Depends(get_db)):
    r = db.query(LossReasonDB).filter(LossReasonDB.reason_id == reason_id).first()
    if not r: return {"error": "Не найдено"}
    db.delete(r); db.commit()
    return {"message": "Удалено"}

# ─────────── PIPELINES ───────────

@app.get("/pipelines")
def list_pipelines(db: Session = Depends(get_db)):
    rows = db.query(PipelineDB).order_by(PipelineDB.sort, PipelineDB.pipeline_id).all()
    return [{
        "pipeline_id": p.pipeline_id, "name": p.name, "client_type": p.client_type,
        "description": p.description, "stages": p.stages.split("|"), "sort": p.sort,
    } for p in rows]

@app.post("/pipelines")
def create_pipeline(p: Pipeline, db: Session = Depends(get_db)):
    if db.query(PipelineDB).filter(PipelineDB.pipeline_id == p.pipeline_id).first():
        return {"error": "ID занят"}
    db.add(PipelineDB(**p.model_dump())); db.commit()
    return {"message": "Воронка создана"}

@app.put("/pipelines/{pipeline_id}")
def update_pipeline(pipeline_id: int, p: Pipeline, db: Session = Depends(get_db)):
    obj = db.query(PipelineDB).filter(PipelineDB.pipeline_id == pipeline_id).first()
    if not obj: return {"error": "Не найдено"}
    obj.name = p.name; obj.client_type = p.client_type
    obj.description = p.description; obj.stages = p.stages; obj.sort = p.sort
    db.commit()
    return {"message": "Сохранено"}

# ─────────── MANAGERS ───────────

def _manager_to_dict(m: ManagerDB) -> dict:
    return {
        "manager_id": m.manager_id, "name": m.name, "phone": m.phone, "email": m.email,
        "role": m.role, "active": m.active,
        "online": _is_online(m),
        "last_seen": m.last_seen.isoformat() if m.last_seen else None,
        "telegram_linked": bool(m.telegram_chat_id),
        "login": m.login,
    }

@app.get("/managers")
def list_managers(db: Session = Depends(get_db), include_inactive: bool = False):
    q = db.query(ManagerDB)
    if not include_inactive:
        q = q.filter(ManagerDB.active == True)
    managers = q.order_by(ManagerDB.manager_id).all()
    return [_manager_to_dict(m) for m in managers]

@app.post("/managers")
def create_manager(m: Manager, db: Session = Depends(get_db)):
    if db.query(ManagerDB).filter(ManagerDB.manager_id == m.manager_id).first():
        return {"error": "ID занят"}
    if m.login:
        if db.query(ManagerDB).filter(ManagerDB.login == m.login).first():
            return {"error": "Такой логин уже занят"}
    else:
        return {"error": "Укажите логин"}
    if not m.password:
        return {"error": "Укажите пароль"}
    db.add(ManagerDB(**m.model_dump())); db.commit()
    return {"message": "Менеджер создан"}

@app.post("/auth/login")
def auth_login(payload: LoginRequest, db: Session = Depends(get_db)):
    """Простая авторизация по логину/паролю. Роль определяется учётной записью,
    так что менеджерам и клинерам не нужно её выбирать отдельно — она приходит в ответе."""
    login = (payload.login or "").strip()
    m = db.query(ManagerDB).filter(ManagerDB.login == login).first()
    if not m or not m.active or not m.password or m.password != payload.password:
        return JSONResponse(status_code=401, content={"error": "Неверный логин или пароль"})
    return _manager_to_dict(m)

@app.post("/managers/{manager_id}/heartbeat")
def manager_heartbeat(manager_id: int, db: Session = Depends(get_db)):
    """Вызывается фронтендом каждые ~30 сек, пока пользователь находится в CRM —
    так мы получаем реальный статус онлайн/оффлайн."""
    m = db.query(ManagerDB).filter(ManagerDB.manager_id == manager_id).first()
    if not m:
        return JSONResponse(status_code=404, content={"error": "Менеджер не найден"})
    m.last_seen = datetime.utcnow()
    db.commit()
    return {"online": True}

@app.put("/managers/{manager_id}")
def update_manager(manager_id: int, payload: ManagerUpdate, db: Session = Depends(get_db)):
    m = db.query(ManagerDB).filter(ManagerDB.manager_id == manager_id).first()
    if not m: return {"error": "Не найден"}
    if payload.login is not None and payload.login != m.login:
        if db.query(ManagerDB).filter(ManagerDB.login == payload.login, ManagerDB.manager_id != manager_id).first():
            return {"error": "Такой логин уже занят"}
    for f in ("name", "phone", "email", "role", "login", "password"):
        v = getattr(payload, f)
        if v is not None: setattr(m, f, _clean(v) or v)
    if payload.active is not None: m.active = payload.active
    db.commit()
    return {"message": "Сохранено"}

@app.get("/managers/{manager_id}/telegram-link")
def get_telegram_link(manager_id: int, db: Session = Depends(get_db)):
    """Персональная ссылка для привязки Telegram клинера/менеджера к его аккаунту."""
    m = db.query(ManagerDB).filter(ManagerDB.manager_id == manager_id).first()
    if not m:
        return JSONResponse(status_code=404, content={"error": "Менеджер не найден"})
    tg_token = _tg_token()
    if not tg_token:
        return {"error": "Telegram-бот не настроен (нет TELEGRAM_BOT_TOKEN)"}
    username = None
    try:
        r = httpx.get(f"https://api.telegram.org/bot{tg_token}/getMe", timeout=8)
        data = r.json()
        if data.get("ok"):
            username = data["result"]["username"]
    except Exception as e:
        print(f"[TG getMe] {e}")
    if not username:
        return {"error": "Не удалось получить данные бота"}
    return {"username": username, "link": f"https://t.me/{username}?start={manager_id}",
            "linked": bool(m.telegram_chat_id)}

@app.put("/leads/assign-manager")
def assign_manager(phone: str, manager_id: Optional[int] = None, db: Session = Depends(get_db)):
    lead = db.query(LeadDB).filter(LeadDB.phone == phone).first()
    if not lead: return {"error": "Лид не найден"}
    if manager_id is not None and not db.query(ManagerDB).filter(ManagerDB.manager_id == manager_id).first():
        return {"error": "Менеджер не найден"}
    lead.manager_id = manager_id
    _log_activity(db, phone, "assigned", {"manager_id": manager_id})
    db.commit()
    return {"message": "Назначен"}

# ─────────── EMAIL ACCOUNTS (IMAP) ───────────

@app.get("/email-accounts")
def list_email_accounts(db: Session = Depends(get_db)):
    return [_email_account_to_dict(a) for a in db.query(EmailAccountDB).order_by(EmailAccountDB.id).all()]

@app.post("/email-accounts")
def create_email_account(payload: EmailAccountIn, db: Session = Depends(get_db)):
    if not db.query(PipelineDB).filter(PipelineDB.pipeline_id == payload.pipeline_id).first():
        return JSONResponse(status_code=422, content={"error": f"Воронка с id {payload.pipeline_id} не найдена"})
    a = EmailAccountDB(
        name=payload.name, email=payload.email, password=payload.password,
        imap_host=payload.imap_host, imap_port=payload.imap_port, imap_ssl=payload.imap_ssl,
        smtp_host=payload.smtp_host, smtp_port=payload.smtp_port, smtp_ssl=payload.smtp_ssl,
        pipeline_id=payload.pipeline_id, client_type=payload.client_type,
        active=payload.active, last_uid=None,
    )
    db.add(a); db.commit(); db.refresh(a)
    if a.active:
        threading.Thread(target=_ensure_idle_threads, daemon=True).start()
    return {"message": "Почтовый ящик добавлен", "data": _email_account_to_dict(a)}

@app.put("/email-accounts/{account_id}")
def update_email_account(account_id: int, payload: EmailAccountUpdate, db: Session = Depends(get_db)):
    a = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
    if not a:
        return JSONResponse(status_code=404, content={"error": "Не найден"})
    data = payload.model_dump(exclude_unset=True)
    # если пришёл пустой пароль — не затираем существующий (значит поле просто не меняли)
    if "password" in data and not data["password"]:
        data.pop("password")
    for k, v in data.items():
        setattr(a, k, v)
    db.commit()
    # если ящик активирован — запускаем IDLE-поток немедленно
    threading.Thread(target=_ensure_idle_threads, daemon=True).start()
    return {"message": "Сохранено", "data": _email_account_to_dict(a)}

@app.delete("/email-accounts/{account_id}")
def delete_email_account(account_id: int, db: Session = Depends(get_db)):
    a = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
    if not a:
        return {"error": "Не найден"}
    db.delete(a); db.commit()
    return {"message": "Ящик удалён"}

@app.post("/email-accounts/{account_id}/test")
def test_email_account(account_id: int, db: Session = Depends(get_db)):
    a = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
    if not a:
        return JSONResponse(status_code=404, content={"error": "Не найден"})
    if not a.password:
        return {"ok": False, "error": "Не указан пароль"}
    ok, err = _email_account_test(a)
    return {"ok": ok, "error": err}

@app.post("/email-accounts/{account_id}/poll")
def poll_email_account_now(account_id: int, db: Session = Depends(get_db)):
    """Ручная проверка почты прямо сейчас (не дожидаясь фонового цикла)."""
    a = db.query(EmailAccountDB).filter(EmailAccountDB.id == account_id).first()
    if not a:
        return JSONResponse(status_code=404, content={"error": "Не найден"})
    if not a.active:
        return {"error": "Ящик выключен"}
    if not a.password:
        return {"error": "Не указан пароль"}
    created = _poll_email_account(account_id)
    db.refresh(a)
    return {"message": f"Проверено. Новых лидов: {created}", "created": created, "data": _email_account_to_dict(a)}

# ─────────── ИМПОРТ ЛИДОВ ИЗ EXCEL ───────────

IMPORT_HEADER_MAP = {
    "имя": "name", "фио": "name", "название": "name", "наименование": "name",
    "имя клиента": "name", "клиент": "name", "name": "name",
    "телефон": "phone", "номер телефона": "phone", "номер": "phone", "phone": "phone", "тел": "phone",
    "email": "email", "почта": "email", "e-mail": "email", "электронная почта": "email",
    "компания": "company_name", "company": "company_name", "организация": "company_name",
    "источник": "source", "source": "source",
    "комментарий": "comment", "примечание": "comment",
}

def _norm_phone(raw) -> Optional[str]:
    if raw is None:
        return None
    s = re.sub(r"[^\d+]", "", str(raw).strip())
    if not s:
        return None
    if s.startswith("8") and len(s) == 11:
        s = "+7" + s[1:]
    elif s.startswith("7") and len(s) == 11:
        s = "+" + s
    elif not s.startswith("+"):
        s = "+" + s
    return s

def _norm_email(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or "@" not in s:
        return None
    return s

@app.get("/leads/import/template")
def download_import_template():
    """Отдаёт xlsx-шаблон с нужными колонками для импорта."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"
    ws.append(["Имя", "Телефон", "Email", "Компания", "Источник", "Комментарий"])
    ws.append(["Иван Иванов", "+79991234567", "ivan@example.com", "ООО «Ромашка»", "Импорт", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="leads_import_template.xlsx"'},
    )

@app.post("/leads/import")
async def import_leads(
    file: UploadFile = File(...),
    pipeline_id: int = Form(1),
    client_type: str = Form("физ"),
    manager_id: Optional[int] = Form(None),
    source: Optional[str] = Form("Импорт из Excel"),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return JSONResponse(status_code=400, content={"error": "Поддерживаются только файлы .xlsx"})

    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == pipeline_id).first()
    if not pipeline:
        return JSONResponse(status_code=400, content={"error": f"Воронка с id {pipeline_id} не найдена"})
    default_status = pipeline.stages.split("|")[0]

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": f"Не удалось прочитать файл: {e}"})
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "skipped": 0, "errors": ["Файл пуст"]}

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_map = {}   # индекс колонки -> поле лида
    for idx, h in enumerate(header):
        field = IMPORT_HEADER_MAP.get(h)
        if field:
            col_map[idx] = field

    if "phone" not in col_map.values():
        return JSONResponse(status_code=400, content={
            "error": "Не найдена колонка с телефоном. Ожидаются заголовки вида «Имя», «Телефон», «Email» "
                     "(см. /leads/import/template)."
        })

    created, skipped, errors = 0, 0, []
    seen_phones_in_file = set()

    for row_num, row in enumerate(rows[1:], start=2):
        data = {}
        for idx, field in col_map.items():
            if idx < len(row):
                data[field] = row[idx]

        phone = _norm_phone(data.get("phone"))
        if not phone:
            skipped += 1
            errors.append(f"Строка {row_num}: не указан или некорректен телефон")
            continue
        if phone in seen_phones_in_file:
            skipped += 1
            errors.append(f"Строка {row_num}: дубликат телефона {phone} внутри файла")
            continue
        seen_phones_in_file.add(phone)

        if db.query(LeadDB).filter(LeadDB.phone == phone).first():
            skipped += 1
            errors.append(f"Строка {row_num}: лид с телефоном {phone} уже есть в CRM")
            continue

        name = _clean(data.get("name")) or phone
        email_val = _norm_email(data.get("email"))

        db_lead = LeadDB(
            name=name, phone=phone, status=default_status,
            email=email_val, pipeline_id=pipeline_id, manager_id=manager_id,
            client_type=client_type,
            source=_clean(data.get("source")) or source,
            company_name=_clean(data.get("company_name")),
            comment=_clean(data.get("comment")),
        )
        db.add(db_lead)
        db.flush()
        _log_activity(db, phone, "lead_created", {"name": name, "source": "import"}, manager_id)
        created += 1

    db.commit()
    return {
        "message": f"Импорт завершён: создано {created}, пропущено {skipped}",
        "created": created, "skipped": skipped,
        "errors": errors[:200],   # не заваливаем ответ, если строк с ошибками очень много
    }

# ─────────── ВНУТРЕННЯЯ РАССЫЛКА ───────────

@app.get("/mail-accounts")
def list_mail_accounts(db: Session = Depends(get_db)):
    """Ящики, у которых заполнен SMTP (пригодны как отправитель рассылки)."""
    accounts = db.query(EmailAccountDB).filter(EmailAccountDB.smtp_host.isnot(None)).all()
    return [_email_account_to_dict(a) for a in accounts if a.password]

@app.post("/mailings")
def create_mailing(payload: MailingCreate, db: Session = Depends(get_db)):
    account = db.query(EmailAccountDB).filter(EmailAccountDB.id == payload.email_account_id).first()
    if not account:
        return JSONResponse(status_code=400, content={"error": "Ящик-отправитель не найден"})
    if not account.smtp_host or not account.password:
        return JSONResponse(status_code=400, content={"error": "У выбранного ящика не настроен SMTP или не указан пароль"})

    q = db.query(LeadDB).filter(LeadDB.email.isnot(None), LeadDB.email != "")
    if payload.lead_ids:
        q = q.filter(LeadDB.id.in_(payload.lead_ids))
    else:
        if payload.pipeline_id:
            q = q.filter(LeadDB.pipeline_id == payload.pipeline_id)
        if payload.tag_id:
            q = q.join(LeadTagDB, LeadTagDB.lead_id == LeadDB.id).filter(LeadTagDB.tag_id == payload.tag_id)
    leads = q.all()

    if not leads:
        return JSONResponse(status_code=400, content={"error": "Не найдено ни одного лида с заполненным email по заданным условиям"})

    campaign_id = uuid.uuid4().hex[:12]
    queued = 0
    for lead in leads:
        db.add(MailQueueDB(
            campaign_id=campaign_id,
            campaign_name=_clean(payload.campaign_name) or payload.subject,
            email_account_id=account.id,
            lead_id=lead.id,
            to_email=lead.email,
            to_name=lead.name,
            subject=_render_template(payload.subject, lead.name),
            body=_render_template(payload.body, lead.name),
            status="pending",
        ))
        queued += 1
    db.commit()
    # отправляем прямо сейчас, синхронно, в этом же запросе — никакой очереди
    # и никакого фонового воркера, письма уходят через SMTP немедленно
    result = _send_campaign_sync(db, campaign_id)
    return {
        "message": f"Рассылка завершена: отправлено {result['sent']}, ошибок {result['failed']}",
        "campaign_id": campaign_id, "queued": queued,
        "sent": result["sent"], "failed": result["failed"],
    }

@app.get("/mailings")
def list_mailings(db: Session = Depends(get_db)):
    """Список рассылок (кампаний) со сводной статистикой по статусам."""
    rows = db.query(
        MailQueueDB.campaign_id, MailQueueDB.campaign_name,
        func.min(MailQueueDB.created_at).label("created_at"),
        func.count(MailQueueDB.id).label("total"),
        func.sum(case((MailQueueDB.status == "sent", 1), else_=0)).label("sent"),
        func.sum(case((MailQueueDB.status == "failed", 1), else_=0)).label("failed"),
        func.sum(case((MailQueueDB.status == "pending", 1), else_=0)).label("pending"),
    ).group_by(MailQueueDB.campaign_id).order_by(func.min(MailQueueDB.created_at).desc()).all()
    return [
        {
            "campaign_id": r.campaign_id, "campaign_name": r.campaign_name,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "total": r.total, "sent": r.sent or 0, "failed": r.failed or 0, "pending": r.pending or 0,
        } for r in rows
    ]

@app.get("/mailings/{campaign_id}")
def mailing_detail(campaign_id: str, db: Session = Depends(get_db)):
    items = db.query(MailQueueDB).filter(MailQueueDB.campaign_id == campaign_id).order_by(MailQueueDB.id).all()
    if not items:
        return JSONResponse(status_code=404, content={"error": "Рассылка не найдена"})
    return [
        {
            "id": i.id, "to_email": i.to_email, "to_name": i.to_name,
            "status": i.status, "attempts": i.attempts, "error": i.error,
            "sent_at": i.sent_at.isoformat() if i.sent_at else None,
        } for i in items
    ]

@app.delete("/mailings/{campaign_id}")
def cancel_mailing(campaign_id: str, db: Session = Depends(get_db)):
    """Отменяет ещё не отправленные письма кампании (уже отправленные не трогает)."""
    n = db.query(MailQueueDB).filter(
        MailQueueDB.campaign_id == campaign_id, MailQueueDB.status == "pending"
    ).update({"status": "failed", "error": "Отменено пользователем"})
    db.commit()
    return {"message": f"Отменено писем: {n}"}

# ─────────── EXPORT ───────────

@app.get("/export/leads.csv")
def export_csv(db: Session = Depends(get_db)):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Имя","Телефон","WhatsApp","Email","Telegram","Viber","VK",
                "Тип","Компания","Адрес","Бюджет","Воронка","Статус","Менеджер","Источник","Теги"])
    leads = db.query(LeadDB).all()
    managers = {m.manager_id: m.name for m in db.query(ManagerDB).all()}
    pipelines = {p.pipeline_id: p.name for p in db.query(PipelineDB).all()}
    for l in leads:
        tags = ", ".join(t.name for t in db.query(TagDB).join(LeadTagDB).filter(LeadTagDB.lead_id == l.id).all())
        w.writerow([l.name, l.phone, l.whatsapp, l.email, l.telegram, l.viber, l.vk,
                    l.client_type, l.company_name, l.address, l.budget,
                    pipelines.get(l.pipeline_id, ""), l.status,
                    managers.get(l.manager_id, ""), l.source, tags])
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="cleanfloor_leads.csv"'})

# ─────────── SETTINGS ───────────

@app.get("/settings/widget-token")
def get_widget_token(db: Session = Depends(get_db)):
    s = db.get(SettingDB, "widget_token")
    return {"token": s.value if s else None}

@app.post("/settings/widget-token/regenerate")
def regenerate_token(db: Session = Depends(get_db)):
    s = db.get(SettingDB, "widget_token")
    new = "cfcrm_" + secrets.token_urlsafe(24)
    if s: s.value = new
    else: db.add(SettingDB(key="widget_token", value=new))
    db.commit()
    return {"token": new}

class TelegramSettings(BaseModel):
    token:   Optional[str] = None
    chat_id: Optional[str] = None

def _set_setting(db: Session, key: str, value: Optional[str]):
    s = db.get(SettingDB, key)
    if s: s.value = value
    else: db.add(SettingDB(key=key, value=value))

@app.get("/settings/telegram")
def get_telegram_settings(db: Session = Depends(get_db)):
    token = db.get(SettingDB, "telegram_bot_token")
    chat_id = db.get(SettingDB, "telegram_notify_chat_id")
    return {"token": token.value if token else "", "chat_id": chat_id.value if chat_id else ""}

@app.post("/settings/telegram")
def save_telegram_settings(payload: TelegramSettings, db: Session = Depends(get_db)):
    if payload.token is not None:
        _set_setting(db, "telegram_bot_token", payload.token.strip() or None)
    if payload.chat_id is not None:
        _set_setting(db, "telegram_notify_chat_id", payload.chat_id.strip() or None)
    db.commit()
    # если задан PUBLIC_URL — сразу переподписываем вебхук на новый токен
    public_url = os.getenv("PUBLIC_URL", "")
    tg_token = _tg_token()
    if tg_token and public_url:
        try:
            httpx.post(f"https://api.telegram.org/bot{tg_token}/setWebhook",
                       json={"url": public_url.rstrip("/") + "/telegram/webhook"}, timeout=8)
        except Exception:
            pass
    return {"message": "Сохранено"}

@app.post("/settings/telegram/detect-chat")
def detect_telegram_chat(db: Session = Depends(get_db)):
    """Находит chat_id по последнему сообщению, отправленному боту (getUpdates).
    Перед вызовом пользователь должен написать боту в личку или в группу."""
    tg_token = _tg_token()
    if not tg_token:
        return {"error": "Сначала укажите и сохраните токен бота"}
    try:
        r = httpx.get(f"https://api.telegram.org/bot{tg_token}/getUpdates", timeout=8)
        data = r.json()
    except Exception as e:
        return {"error": f"Не удалось связаться с Telegram: {e}"}
    if not data.get("ok"):
        return {"error": data.get("description", "Ошибка Telegram API")}
    updates = data.get("result", [])
    if not updates:
        return {"error": "Нет новых сообщений боту. Напишите боту что-нибудь в Telegram и попробуйте снова."}
    last = updates[-1]
    msg = last.get("message") or last.get("channel_post") or {}
    chat = msg.get("chat", {})
    chat_id = chat.get("id")
    if not chat_id:
        return {"error": "Не удалось определить chat_id из последнего сообщения"}
    _set_setting(db, "telegram_notify_chat_id", str(chat_id))
    db.commit()
    title = chat.get("title") or " ".join(filter(None, [chat.get("first_name"), chat.get("last_name")])) or chat.get("username") or str(chat_id)
    return {"chat_id": str(chat_id), "title": title}

@app.post("/settings/telegram/test")
def test_telegram_notification(db: Session = Depends(get_db)):
    chat_id = _tg_notify_chat_id()
    if not _tg_token():
        return {"error": "Сначала укажите токен бота"}
    if not chat_id:
        return {"error": "Сначала укажите или определите chat_id"}
    res = _tg_send(chat_id, "🆕 *Новая заявка*\n👤 Имя: Егор\n📞 Телефон: +49838\n📍 Откуда: Тест из CRM")
    if res and res.get("ok"):
        return {"message": "Отправлено"}
    return {"error": (res or {}).get("description", "Не удалось отправить сообщение")}

# ─────────── PUBLIC ───────────

class PublicLead(BaseModel):
    name:        str
    phone:       str
    pipeline_id: int = 1
    client_type: str = "физ"
    telegram:    Optional[str] = None
    whatsapp:    Optional[str] = None
    source:      Optional[str] = None
    message:     Optional[str] = None

def _widget_token(db):
    s = db.get(SettingDB, "widget_token")
    return s.value if s else ""

@app.post("/public/leads")
async def public_create(request: Request, db: Session = Depends(get_db)):
    token = request.headers.get("X-CRM-Token", "")
    if not token or token != _widget_token(db):
        return JSONResponse(status_code=403, content={"error": "Неверный токен"})
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=422, content={"error": "Неверный JSON"})
    name        = (body.get("name") or "").strip()
    phone       = (body.get("phone") or "").strip()
    pipeline_id = int(body.get("pipeline_id", 1))
    client_type = body.get("client_type", "физ")
    telegram    = body.get("telegram") or None
    whatsapp    = body.get("whatsapp") or None
    source      = body.get("source") or body.get("button") or None
    source      = source.strip() if isinstance(source, str) else source
    if not name or not phone:
        return JSONResponse(status_code=422, content={"error": "name и phone обязательны"})
    if db.query(LeadDB).filter(LeadDB.phone == phone).first():
        return {"message": "Лид уже существует", "phone": phone}
    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == pipeline_id).first()
    status = pipeline.stages.split("|")[0] if pipeline else "Новая заявка"
    lead = LeadDB(name=name, phone=phone, status=status, pipeline_id=pipeline_id,
                  client_type=client_type, telegram=telegram, whatsapp=whatsapp, source=source)
    db.add(lead); db.commit(); db.refresh(lead)
    msg = body.get("message")
    if msg:
        db.add(NoteDB(phone=phone, text=f"Сообщение с формы: {msg}", note_type="system"))
        db.commit()
    _log_activity(db, phone, "lead_created", {"source": source, "public": True})
    db.commit()
    send_fcm_push(name, phone)
    _notify_new_lead(name, phone, source)
    return {"message": "Заявка принята", "status": status}

# ─────────── TELEGRAM-БОТ ДЛЯ КЛИЕНТОВ: СЦЕНАРИЙ ───────────
# Полный клиентский сценарий (как в отдельном боте cleanfloor_bot.py), встроенный
# прямо в вебхук CRM — работает на том же токене/боте, что и уведомления клинерам.
#
# /start -> 3 кнопки:
#   🧹 Заказать уборку        -> подробная анкета -> лид в CRM (пайплайн физ/юр по типу объекта)
#   📞 Заказать звонок        -> имя, фамилия, телефон -> лид в CRM
#   💬 Связаться с менеджером -> свободный вопрос -> уходит в чат уведомлений (Настройки -> Telegram),
#                                 менеджер отвечает Reply-сообщением, ответ уходит клиенту

BOT_SOURCE_CLEANING = "Telegram-бот — Заказать уборку"
BOT_SOURCE_CALL = "Telegram-бот — Заказать звонок"

# Состояние диалога на чат: {chat_id: {"state": "...", "data": {...}}}
# Хранится в памяти процесса (как MemoryStorage у aiogram) — сценарий начинается заново после рестарта.
_bot_sessions: dict = {}

# message_id уведомления в чате менеджера -> chat_id клиента, задавшего вопрос
_pending_questions: dict = {}

BOT_OBJECT_TYPES = {
    "obj:flat": "Квартира",
    "obj:house": "Дом",
    "obj:office": "Офис",
    "obj:shop": "Магазин",
    "obj:industrial": "Производственное помещение",
}
BOT_OBJECT_TO_CLIENT_TYPE = {
    "obj:flat": "физ", "obj:house": "физ",
    "obj:office": "юр", "obj:shop": "юр", "obj:industrial": "юр",
}
BOT_CLEANING_TYPES = {
    "clean:support": "Поддерживающая уборка",
    "clean:general": "Генеральная уборка",
    "clean:after_repair": "После ремонта",
    "clean:windows": "Мытье окон",
    "clean:furniture": "Химчистка мебели",
    "clean:other": "Другое",
}
BOT_AREAS = {
    "area:lt50": "До 50 м²", "area:50_100": "50–100 м²",
    "area:100_200": "100–200 м²", "area:gt200": "Более 200 м²",
}
BOT_WHEN_OPTIONS = {"when:today": "Сегодня", "when:tomorrow": "Завтра", "when:week": "На этой неделе"}

def _bot_kb(options: dict) -> dict:
    return {"inline_keyboard": [[{"text": text, "callback_data": key}] for key, text in options.items()]}

def _bot_main_menu_kb() -> dict:
    return {"inline_keyboard": [
        [{"text": "🧹 Заказать уборку", "callback_data": "menu:cleaning"}],
        [{"text": "📞 Заказать звонок", "callback_data": "menu:call"}],
        [{"text": "💬 Связаться с менеджером", "callback_data": "menu:manager"}],
    ]}

def _bot_phone_kb() -> dict:
    return {"keyboard": [[{"text": "📱 Отправить номер телефона", "request_contact": True}]],
            "resize_keyboard": True, "one_time_keyboard": True}

_BOT_REMOVE_KB = {"remove_keyboard": True}

def _bot_is_valid_phone(txt: str) -> bool:
    cleaned = re.sub(r"[^\d+]", "", txt or "")
    return bool(re.fullmatch(r"(\+?375\d{9}|\+?7\d{10}|8\d{10})", cleaned))

def _bot_get_session(chat_id: str) -> dict:
    return _bot_sessions.setdefault(chat_id, {"state": None, "data": {}})

def _bot_clear_session(chat_id: str):
    _bot_sessions.pop(chat_id, None)

def _bot_create_lead(db: Session, name: str, phone: str, source: str,
                      client_type: str = "физ", comment: Optional[str] = None):
    """Создаёт лид из бота-воронки (физ -> pipeline 1, юр -> pipeline 2).
    Если номер уже есть в CRM — не дублирует лид, а добавляет заметку о повторном обращении."""
    pipeline_id = 2 if client_type == "юр" else 1
    existing = db.query(LeadDB).filter(LeadDB.phone == phone).first()
    if existing:
        note = f"Повторное обращение через Telegram-бота ({source})" + (f": {comment}" if comment else "")
        db.add(NoteDB(phone=phone, text=note, note_type="system"))
        _log_activity(db, phone, "note_added", {"source": source})
        db.commit()
        return existing
    pipeline = db.query(PipelineDB).filter(PipelineDB.pipeline_id == pipeline_id).first()
    status = pipeline.stages.split("|")[0] if pipeline else "Новый лид"
    lead = LeadDB(name=name, phone=phone, status=status, pipeline_id=pipeline_id,
                  client_type=client_type, source=source, comment=comment)
    db.add(lead); db.commit(); db.refresh(lead)
    _log_activity(db, phone, "lead_created", {"name": name, "source": source})
    db.commit()
    return lead

async def _handle_client_bot_update(update: dict, db: Session) -> bool:
    """Обрабатывает апдейт клиентского сценария бота. Возвращает True, если апдейт обработан."""
    msg = update.get("message")
    if msg:
        chat_id = str(msg.get("chat", {}).get("id", ""))
        text_in = (msg.get("text") or "").strip()
        contact = msg.get("contact")
        notify_chat = _tg_notify_chat_id()

        # Ответ менеджера клиенту: менеджер отвечает (Reply) в чате уведомлений на вопрос клиента
        if notify_chat and chat_id == str(notify_chat) and msg.get("reply_to_message"):
            original_id = msg["reply_to_message"].get("message_id")
            client_chat_id = _pending_questions.get(original_id)
            if client_chat_id and text_in:
                _tg_send(client_chat_id, f"Ответ менеджера: {text_in}")
                _tg_send(chat_id, "✅ Ответ отправлен клиенту.")
                _pending_questions.pop(original_id, None)
                return True

        if text_in.startswith("/start"):
            # /start <manager_id> обрабатывается отдельно (привязка аккаунта менеджера/клинера) — см. ниже
            parts = text_in.split(maxsplit=1)
            if len(parts) == 2 and parts[1].strip().isdigit():
                return False
            _bot_clear_session(chat_id)
            _tg_send(chat_id,
                     "Здравствуйте! Мы профессионально выполняем уборку квартир, домов, офисов "
                     "и коммерческих помещений.\n\n"
                     "✅ Работаем без выходных\n"
                     "✅ Используем профессиональную химию и оборудование\n"
                     "✅ Гарантируем качество уборки\n\n"
                     "Выберите, что вас интересует:",
                     reply_markup=_bot_main_menu_kb())
            return True

        session = _bot_sessions.get(chat_id)
        state = session.get("state") if session else None
        if not state:
            return False  # не в клиентском сценарии — не обрабатываем (могло быть иное сообщение)
        data = session["data"]

        if state == "cleaning:cleaning_type_custom":
            data["cleaning_type"] = text_in
            session["state"] = "cleaning:area"
            _tg_send(chat_id, "Какая площадь помещения?", reply_markup=_bot_kb(BOT_AREAS))
            return True

        if state == "cleaning:address":
            data["address"] = text_in
            session["state"] = "cleaning:phone"
            _tg_send(chat_id, "Оставьте номер телефона (или отправьте его кнопкой ниже):",
                     reply_markup=_bot_phone_kb())
            return True

        if state == "cleaning:phone":
            phone = contact.get("phone_number") if contact else text_in
            if not contact and not _bot_is_valid_phone(phone):
                _tg_send(chat_id, "Похоже, это не номер телефона. Введите номер полностью, "
                                   "например: +375291234567, или отправьте его кнопкой ниже.",
                         reply_markup=_bot_phone_kb())
                return True
            data["phone"] = phone
            session["state"] = "cleaning:name"
            _tg_send(chat_id, "Оставьте имя и фамилию:", reply_markup=_BOT_REMOVE_KB)
            return True

        if state == "cleaning:name":
            full_name = text_in
            client_type = data.get("client_type", "физ")
            comment = (
                f"Объект: {data.get('object_type')}; Вид уборки: {data.get('cleaning_type')}; "
                f"Площадь: {data.get('area')}; Когда: {data.get('when')}; Адрес/район: {data.get('address')}"
            )
            _bot_create_lead(db, full_name, data.get("phone", ""), BOT_SOURCE_CLEANING,
                              client_type=client_type, comment=comment)
            _notify_new_lead(full_name, data.get("phone", ""), BOT_SOURCE_CLEANING, comment)
            _tg_send(chat_id, "Спасибо! Ваша заявка принята. В ближайшее время с вами свяжется менеджер.")
            _tg_send(chat_id, "Чтобы вернуться в меню — введите /start")
            _bot_clear_session(chat_id)
            return True

        if state == "call:name":
            data["name"] = text_in
            session["state"] = "call:surname"
            _tg_send(chat_id, "Введите вашу фамилию:")
            return True

        if state == "call:surname":
            data["surname"] = text_in
            session["state"] = "call:phone"
            _tg_send(chat_id, "Введите ваш номер телефона (или отправьте его кнопкой ниже):",
                     reply_markup=_bot_phone_kb())
            return True

        if state == "call:phone":
            phone = contact.get("phone_number") if contact else text_in
            if not contact and not _bot_is_valid_phone(phone):
                _tg_send(chat_id, "Похоже, это не номер телефона. Введите номер полностью, "
                                   "например: +375291234567, или отправьте его кнопкой ниже.",
                         reply_markup=_bot_phone_kb())
                return True
            full_name = f"{data.get('name', '')} {data.get('surname', '')}".strip()
            _bot_create_lead(db, full_name, phone, BOT_SOURCE_CALL)
            _notify_new_lead(full_name, phone, BOT_SOURCE_CALL)
            _tg_send(chat_id, "Спасибо! Ваша заявка принята. В ближайшее время с вами свяжется менеджер.",
                     reply_markup=_BOT_REMOVE_KB)
            _tg_send(chat_id, "Чтобы вернуться в меню — введите /start")
            _bot_clear_session(chat_id)
            return True

        if state == "question:question":
            _bot_clear_session(chat_id)
            if notify_chat:
                res = _tg_send(notify_chat, f"❗️Вопрос от клиента (chat {chat_id})\n\n{text_in}\n\n"
                                              "👉 Чтобы ответить клиенту — ответьте (Reply) на это сообщение.")
                if res and res.get("ok"):
                    _pending_questions[res["result"]["message_id"]] = chat_id
            _tg_send(chat_id, "Спасибо за ваш вопрос! Мы ответим вам как можно скорее.")
            return True

        return False

    cq = update.get("callback_query")
    if cq:
        data_str = cq.get("data", "")
        cq_id = cq.get("id")
        chat_id = str(cq.get("message", {}).get("chat", {}).get("id", ""))

        if data_str == "menu:cleaning":
            session = _bot_get_session(chat_id); session["state"] = "cleaning:object_type"; session["data"] = {}
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Какой объект необходимо убрать?", reply_markup=_bot_kb(BOT_OBJECT_TYPES))
            return True

        if data_str == "menu:call":
            session = _bot_get_session(chat_id); session["state"] = "call:name"; session["data"] = {}
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Введите ваше имя:")
            return True

        if data_str == "menu:manager":
            session = _bot_get_session(chat_id); session["state"] = "question:question"; session["data"] = {}
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Напишите ваш вопрос, а мы ответим на него как можно быстрее.")
            return True

        session = _bot_sessions.get(chat_id)
        state = session.get("state") if session else None

        if state == "cleaning:object_type" and data_str.startswith("obj:") and data_str in BOT_OBJECT_TYPES:
            session["data"]["object_type"] = BOT_OBJECT_TYPES[data_str]
            session["data"]["client_type"] = BOT_OBJECT_TO_CLIENT_TYPE[data_str]
            session["state"] = "cleaning:cleaning_type"
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Какой вид уборки нужен?", reply_markup=_bot_kb(BOT_CLEANING_TYPES))
            return True

        if state == "cleaning:cleaning_type" and data_str.startswith("clean:") and data_str in BOT_CLEANING_TYPES:
            _tg_answer_callback(cq_id, "")
            if data_str == "clean:other":
                session["state"] = "cleaning:cleaning_type_custom"
                _tg_send(chat_id, "Опишите, какой вид уборки вам нужен:")
                return True
            session["data"]["cleaning_type"] = BOT_CLEANING_TYPES[data_str]
            session["state"] = "cleaning:area"
            _tg_send(chat_id, "Какая площадь помещения?", reply_markup=_bot_kb(BOT_AREAS))
            return True

        if state == "cleaning:area" and data_str.startswith("area:") and data_str in BOT_AREAS:
            session["data"]["area"] = BOT_AREAS[data_str]
            session["state"] = "cleaning:when"
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Когда нужна уборка?", reply_markup=_bot_kb(BOT_WHEN_OPTIONS))
            return True

        if state == "cleaning:when" and data_str.startswith("when:") and data_str in BOT_WHEN_OPTIONS:
            session["data"]["when"] = BOT_WHEN_OPTIONS[data_str]
            session["state"] = "cleaning:address"
            _tg_answer_callback(cq_id, "")
            _tg_send(chat_id, "Введите адрес или район:")
            return True

        return False

    return False

# ─────────── TELEGRAM WEBHOOK ───────────

@app.post("/telegram/webhook")
async def telegram_webhook(request: Request, db: Session = Depends(get_db)):
    if not _tg_token():
        return {"ok": True}
    try:
        update = await request.json()
    except Exception:
        return {"ok": True}

    handled = await _handle_client_bot_update(update, db)
    if handled:
        return {"ok": True}

    msg = update.get("message")
    if msg:
        chat_id = str(msg.get("chat", {}).get("id", ""))
        text_in = (msg.get("text") or "").strip()
        if text_in.startswith("/start"):
            parts = text_in.split(maxsplit=1)
            if len(parts) == 2 and parts[1].strip().isdigit():
                mgr = db.query(ManagerDB).filter(ManagerDB.manager_id == int(parts[1].strip())).first()
                if mgr:
                    mgr.telegram_chat_id = chat_id
                    db.commit()
                    _tg_send(chat_id, f"✅ Telegram подключен к аккаунту *{mgr.name}*.\n"
                                       f"Теперь при появлении заявок со статусом «Выезд на объект» "
                                       f"вы будете получать уведомления с кнопкой «Взять в работу».")
                else:
                    _tg_send(chat_id, "Аккаунт не найден. Попросите менеджера прислать актуальную ссылку.")
            else:
                _tg_send(chat_id, "Здравствуйте! Чтобы подключить уведомления, откройте персональную ссылку "
                                   "из CRM (раздел «Менеджеры» → «Подключить Telegram»).")
        return {"ok": True}

    cq = update.get("callback_query")
    if cq:
        data = cq.get("data", "")
        cq_id = cq.get("id")
        from_chat = str(cq.get("message", {}).get("chat", {}).get("id", ""))
        if data.startswith("take:"):
            try:
                lead_id = int(data.split(":", 1)[1])
            except ValueError:
                _tg_answer_callback(cq_id, "Некорректная заявка")
                return {"ok": True}
            lead = db.query(LeadDB).filter(LeadDB.id == lead_id).first()
            cleaner = db.query(ManagerDB).filter(ManagerDB.telegram_chat_id == from_chat).first()
            if not lead:
                _tg_answer_callback(cq_id, "Заявка не найдена", show_alert=True)
            elif lead.status != "Выезд на объект" or lead.taken_by:
                _tg_answer_callback(cq_id, f"Уже взято: {lead.taken_by or '—'}", show_alert=True)
            elif not cleaner:
                _tg_answer_callback(cq_id, "Ваш Telegram не привязан к аккаунту клинера в CRM", show_alert=True)
            else:
                old = lead.status
                lead.status = "Взято в работу"
                lead.taken_by = cleaner.name
                lead.taken_by_manager_id = cleaner.manager_id
                db.add(StageHistoryDB(phone=lead.phone, old_status=old, new_status="Взято в работу", manager_id=cleaner.manager_id))
                _log_activity(db, lead.phone, "stage_change", {"from": old, "to": "Взято в работу", "cleaner": cleaner.name}, cleaner.manager_id)
                db.commit()
                db.refresh(lead)
                _tg_answer_callback(cq_id, "Вы взяли заявку в работу ✅")
                _broadcast_lead_taken(db, lead)
        return {"ok": True}
    return {"ok": True}

@app.post("/telegram/setup-webhook")
def telegram_setup_webhook(base_url: str, db: Session = Depends(get_db)):
    """Разовая настройка: указываете публичный адрес CRM, бот подписывается на вебхук."""
    tg_token = _tg_token()
    if not tg_token:
        return {"error": "TELEGRAM_BOT_TOKEN не задан"}
    url = base_url.rstrip("/") + "/telegram/webhook"
    try:
        r = httpx.post(f"https://api.telegram.org/bot{tg_token}/setWebhook", json={"url": url}, timeout=8)
        return r.json()
    except Exception as e:
        return {"error": str(e)}

# ─────────── WIDGET.JS ───────────

WIDGET_JS = r"""
(function() {
  var scriptEl = document.currentScript;
  var token = scriptEl.getAttribute('data-token');
  var pipelineId = parseInt(scriptEl.getAttribute('data-pipeline') || '1');
  var source = scriptEl.getAttribute('data-source') || null;
  var label = scriptEl.getAttribute('data-label') || 'Оставить заявку';
  var apiBase = scriptEl.src.replace('/widget.js', '');

  window.__cfcrmInstances = (window.__cfcrmInstances || 0);
  var uid = 'cfcrm' + (window.__cfcrmInstances++);
  var bottomOffset = 24 + (window.__cfcrmInstances - 1) * 64;

  if (!document.getElementById('cfcrm-style')) {
    var style = document.createElement('style');
    style.id = 'cfcrm-style';
    style.textContent = [
      '.cfcrm-btn{position:fixed;right:24px;z-index:9999;background:#6D28D9;color:#fff;border:none;border-radius:50px;padding:12px 22px;font-size:14px;font-weight:600;cursor:pointer;box-shadow:0 4px 20px rgba(109,40,217,.4);font-family:inherit;transition:all .2s;}',
      '.cfcrm-btn:hover{background:#8B5CF6;transform:translateY(-2px);}',
      '.cfcrm-overlay{display:none;position:fixed;inset:0;z-index:10000;background:rgba(0,0,0,.6);align-items:center;justify-content:center;padding:20px;}',
      '.cfcrm-overlay.open{display:flex;}',
      '.cfcrm-modal{background:#111118;border:1px solid #242430;border-radius:16px;width:100%;max-width:400px;padding:28px;font-family:inherit;}',
      '.cfcrm-modal h3{color:#F9FAFB;font-size:18px;margin:0 0 6px;}',
      '.cfcrm-modal p{color:#6B7280;font-size:13px;margin:0 0 20px;}',
      '.cfcrm-modal input,.cfcrm-modal textarea{width:100%;padding:10px 14px;background:#18181F;border:1px solid #242430;border-radius:8px;color:#F9FAFB;font-size:14px;margin-bottom:12px;box-sizing:border-box;font-family:inherit;outline:none;}',
      '.cfcrm-modal input:focus,.cfcrm-modal textarea:focus{border-color:#6D28D9;}',
      '.cfcrm-modal button.cfcrm-submit{width:100%;padding:11px;background:#6D28D9;color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;font-family:inherit;}',
      '.cfcrm-modal button.cfcrm-submit:hover{background:#8B5CF6;}',
      '.cfcrm-close{float:right;cursor:pointer;color:#6B7280;font-size:20px;border:none;background:none;line-height:1;}',
      '.cfcrm-ok{display:none;text-align:center;color:#10B981;font-size:15px;padding:10px 0;}'
    ].join('');
    document.head.appendChild(style);
  }

  var btn = document.createElement('button');
  btn.className = 'cfcrm-btn';
  btn.id = uid + '-btn';
  btn.style.bottom = bottomOffset + 'px';
  btn.textContent = '📩 ' + label;
  document.body.appendChild(btn);

  var overlay = document.createElement('div');
  overlay.className = 'cfcrm-overlay';
  overlay.id = uid + '-overlay';
  overlay.innerHTML =
    '<div class="cfcrm-modal">' +
    '<button class="cfcrm-close">✕</button>' +
    '<h3>' + label + '</h3>' +
    '<p>Мы свяжемся с вами в ближайшее время</p>' +
    '<div class="cfcrm-ok">✅ Заявка принята! Мы скоро свяжемся.</div>' +
    '<input class="cfcrm-name" type="text" placeholder="Ваше имя *" />' +
    '<input class="cfcrm-phone" type="tel" placeholder="Телефон *" />' +
    '<input class="cfcrm-tg" type="text" placeholder="Telegram (необязательно)" />' +
    '<textarea class="cfcrm-msg" rows="3" placeholder="Комментарий (необязательно)"></textarea>' +
    '<button type="button" class="cfcrm-submit">Отправить заявку</button>' +
    '</div>';
  document.body.appendChild(overlay);

  var closeBtn = overlay.querySelector('.cfcrm-close');
  var nameInput = overlay.querySelector('.cfcrm-name');
  var phoneInput= overlay.querySelector('.cfcrm-phone');
  var tgInput   = overlay.querySelector('.cfcrm-tg');
  var msgInput  = overlay.querySelector('.cfcrm-msg');
  var submitBtn = overlay.querySelector('.cfcrm-submit');
  var okBox     = overlay.querySelector('.cfcrm-ok');

  btn.onclick = function() { overlay.classList.add('open'); };
  closeBtn.onclick = function() { overlay.classList.remove('open'); };
  overlay.onclick = function(e) { if (e.target === overlay) overlay.classList.remove('open'); };

  submitBtn.onclick = function() {
    var name  = nameInput.value.trim();
    var phone = phoneInput.value.trim();
    var tg    = tgInput.value.trim();
    var msg   = msgInput.value.trim();
    if (!name || !phone) { alert('Заполните имя и телефон'); return; }
    fetch(apiBase + '/public/leads', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json', 'X-CRM-Token': token },
      body: JSON.stringify({ name: name, phone: phone, telegram: tg || null, message: msg || null, pipeline_id: pipelineId, source: source })
    }).then(function(r) { return r.json(); }).then(function(d) {
      okBox.style.display = 'block';
      nameInput.style.display = 'none';
      phoneInput.style.display = 'none';
      tgInput.style.display = 'none';
      msgInput.style.display = 'none';
      submitBtn.style.display = 'none';
      setTimeout(function() {
        overlay.classList.remove('open');
        setTimeout(function() {
          okBox.style.display = 'none';
          nameInput.style.display = '';
          phoneInput.style.display = '';
          tgInput.style.display = '';
          msgInput.style.display = '';
          submitBtn.style.display = '';
          nameInput.value = ''; phoneInput.value = ''; tgInput.value = ''; msgInput.value = '';
        }, 300);
      }, 2500);
    }).catch(function() { alert('Ошибка отправки. Попробуйте позже.'); });
  };
})();
"""

@app.get("/widget.js", include_in_schema=False)
def serve_widget():
    return Response(content=WIDGET_JS, media_type="application/javascript")

# ─────────── STATIC ───────────

STATIC_DIR = os.path.dirname(os.path.abspath(__file__))
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/", include_in_schema=False)
def serve_index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"),
                        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})

@app.get("/favicon.ico", include_in_schema=False)
def serve_favicon():
    p = os.path.join(STATIC_DIR, "favicon.png")
    return FileResponse(p, media_type="image/png") if os.path.exists(p) else JSONResponse(status_code=404, content={"detail": "favicon"})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT", 8000)), reload=False)
